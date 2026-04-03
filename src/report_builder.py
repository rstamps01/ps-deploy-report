"""
VAST As-Built Report Generator - Report Builder Module

This module generates professional PDF reports from processed VAST cluster data.
It creates comprehensive as-built reports with enhanced features including
rack positioning, PSNT tracking, and professional formatting.

Features:
- Professional PDF report generation
- Enhanced data visualization with rack layouts
- PSNT integration for support tracking
- Comprehensive report sections and formatting
- Customizable templates and styling
- Error handling and graceful degradation

Author: Manus AI
Date: September 26, 2025
"""

import json
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

# Add src directory to Python path
sys.path.insert(0, str(Path(__file__).parent))

from brand_compliance import VastBrandCompliance, create_vast_brand_compliance

# Import rack diagram module
from rack_diagram import RackDiagram
from utils import get_bundle_dir, get_data_dir
from utils.logger import get_logger

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, inch
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas
    from reportlab.platypus import (
        BaseDocTemplate,
        Flowable,
        Frame,
        Image,
        KeepTogether,
        NextPageTemplate,
        PageBreak,
        PageTemplate,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


@dataclass
class ReportConfig:
    """Configuration for report generation."""

    page_size: str = "A4"
    margin_top: float = 1.0
    margin_bottom: float = 1.0
    margin_left: float = 1.0
    margin_right: float = 1.0
    font_name: str = "Helvetica"
    font_size: int = 10
    title_font_size: int = 16
    heading_font_size: int = 12
    line_spacing: float = 1.2
    include_toc: bool = True
    include_page_numbers: bool = True
    include_timestamp: bool = True
    include_enhanced_features: bool = True
    organization: str = "VAST Professional Services"
    sections: Dict[str, bool] = field(default_factory=dict)
    network_diagram: Dict[str, Any] = field(default_factory=lambda: {
        "mode": "detailed",
        "show_port_labels": False,
        "device_icons": "flat",
        "orientation": "portrait",
    })

    def section_enabled(self, key: str) -> bool:
        """Return whether a report section should be rendered (default True)."""
        return self.sections.get(key, True)

    @classmethod
    def from_yaml(cls, config: Dict[str, Any]) -> "ReportConfig":
        """Build a ReportConfig from a parsed config.yaml dictionary.

        Supports both flat (``report.margin_top``) and nested
        (``report.template.margin_top``, ``report.pdf.font_size``) layouts.
        """
        report = config.get("report", {})
        template = report.get("template", {})
        pdf = report.get("pdf", {})
        kwargs: Dict[str, Any] = {}

        def _set(field: str, value: Any, conv: type) -> None:
            if value is not None:
                try:
                    kwargs[field] = conv(value)
                except (ValueError, TypeError):
                    pass

        _set("page_size", template.get("page_size", report.get("page_size")), str)
        for m in ("margin_top", "margin_bottom", "margin_left", "margin_right"):
            _set(m, template.get(m, report.get(m)), float)
        _set("font_name", pdf.get("font_family", pdf.get("font_name", report.get("font_name"))), str)
        _set("font_size", pdf.get("font_size", report.get("font_size")), int)
        _set("title_font_size", pdf.get("title_font_size", report.get("title_font_size")), int)
        _set("heading_font_size", pdf.get("heading_font_size", report.get("heading_font_size")), int)
        _set("line_spacing", report.get("line_spacing"), float)
        _set("include_toc", pdf.get("include_toc", report.get("include_toc")), bool)
        _set("include_page_numbers", pdf.get("include_page_numbers", report.get("include_page_numbers")), bool)
        _set("include_timestamp", report.get("include_timestamp"), bool)
        _set("include_enhanced_features", report.get("include_enhanced_features"), bool)
        _set("organization", report.get("organization"), str)

        dc = config.get("data_collection", {})
        raw_sections = dc.get("sections", {})
        if isinstance(raw_sections, dict):
            kwargs["sections"] = {k: bool(v) for k, v in raw_sections.items()}

        nd = config.get("network_diagram", {})
        if isinstance(nd, dict) and nd:
            kwargs["network_diagram"] = nd

        return cls(**kwargs)


@dataclass
class ReportSection:
    """Data class for report section information."""

    title: str
    content: List[Any]
    level: int = 1
    page_break_before: bool = False


class ReportGenerationError(Exception):
    """Custom exception for report generation errors."""

    pass


class PageMarker(Flowable):
    """Invisible flowable that captures page numbers during PDF rendering."""

    def __init__(self, section_key: str, page_tracker: Dict[str, int]):
        """
        Initialize PageMarker.

        Args:
            section_key: Unique identifier for the section (e.g., 'exec_summary')
            page_tracker: Dictionary to store captured page numbers
        """
        self.section_key = section_key
        self.page_tracker = page_tracker
        self.width = 0  # Invisible - doesn't affect layout
        self.height = 0  # Invisible - doesn't affect layout

    def draw(self):
        """Capture current page number during PDF rendering."""
        page_num = self.canv.getPageNumber()
        self.page_tracker[self.section_key] = page_num


class VastReportBuilder:
    """
    VAST As-Built Report Builder for generating professional PDF reports.

    This class creates comprehensive PDF reports from processed VAST cluster data
    with enhanced features for rack positioning and PSNT tracking.
    """

    def __init__(
        self,
        config: Optional[ReportConfig] = None,
        library_path: Optional[str] = None,
        user_images_dir: Optional[str] = None,
    ):
        """
        Initialize the report builder.

        Args:
            config (ReportConfig, optional): Report configuration
            library_path: Path to user device_library.json
            user_images_dir: Path to user-uploaded hardware images directory
        """
        self.logger = get_logger(__name__)
        self.config = config or ReportConfig()
        self.library_path = library_path
        self.user_images_dir = user_images_dir
        self.switch_positions: dict[int, Any] = {}

        if not REPORTLAB_AVAILABLE:
            raise ReportGenerationError("ReportLab is not available. Please install it: pip install reportlab")

        # Initialize VAST brand compliance
        self.brand_compliance = create_vast_brand_compliance()

        self.logger.info("Report builder initialized with VAST brand compliance")

    def _font(self, variant: str = "normal") -> str:
        """Return the ReportLab font name for the configured font family.

        Variant is one of: normal, bold, italic, bold-italic.
        """
        base = self.config.font_name
        _MAP = {
            "Helvetica": {"normal": "Helvetica", "bold": "Helvetica-Bold", "italic": "Helvetica-Oblique", "bold-italic": "Helvetica-BoldOblique"},
            "Times-Roman": {"normal": "Times-Roman", "bold": "Times-Bold", "italic": "Times-Italic", "bold-italic": "Times-BoldItalic"},
            "Courier": {"normal": "Courier", "bold": "Courier-Bold", "italic": "Courier-Oblique", "bold-italic": "Courier-BoldOblique"},
        }
        family = _MAP.get(base, _MAP["Helvetica"])
        return family.get(variant, family["normal"])

    def generate_pdf_report(self, processed_data: Dict[str, Any], output_path: str) -> bool:
        """
        Generate a professional PDF report from processed data.

        Args:
            processed_data (Dict[str, Any]): Processed cluster data
            output_path (str): Output file path for the PDF

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            self.logger.info(f"Generating PDF report: {output_path}")

            # Create output directory if it doesn't exist
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            return self._generate_with_reportlab(processed_data, str(output_file))

        except Exception as e:
            self.logger.error(f"Error generating PDF report: {e}")
            return False

    def _create_landscape_template(
        self, portrait_size: tuple, margins: Dict[str, float]
    ) -> Any:
        """Build a landscape PageTemplate for wide diagrams."""
        from reportlab.platypus import Frame, PageTemplate

        landscape_w = portrait_size[1]
        landscape_h = portrait_size[0]
        left = margins.get("left", 36)
        right = margins.get("right", 36)
        top = margins.get("top", 36)
        bottom = margins.get("bottom", 54)

        frame = Frame(
            left, bottom,
            landscape_w - left - right,
            landscape_h - top - bottom,
            leftPadding=0, bottomPadding=0,
            rightPadding=0, topPadding=0,
        )
        return PageTemplate(
            id="landscape",
            frames=[frame],
            pagesize=(landscape_w, landscape_h),
        )

    def _build_report_story(
        self,
        processed_data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
    ) -> List[Any]:
        """
        Build report story with optional page tracking.

        Args:
            processed_data: Processed cluster data
            page_tracker: Optional dictionary to capture page numbers (for first pass)

        Returns:
            List of flowables for the PDF story
        """
        story = []

        # Determine if we're in first pass (capturing pages) or second pass (using captured pages)
        is_first_pass = page_tracker is not None and len(page_tracker) == 0

        # Add title page
        story.extend(self._create_title_page(processed_data))
        story.append(PageBreak())

        # Add table of contents
        if self.config.include_toc:
            if is_first_pass:
                # First pass: use placeholder TOC (will be replaced in second pass)
                story.extend(self._create_table_of_contents(processed_data))
            else:
                # Second pass: use dynamic TOC with captured page numbers
                story.extend(self._create_table_of_contents_dynamic(processed_data, page_tracker or {}))
            story.append(PageBreak())

        # Add executive summary
        if self.config.section_enabled("executive_summary"):
            story.extend(
                self._create_executive_summary(processed_data, page_tracker, "exec_summary" if is_first_pass else None)
            )
            story.append(PageBreak())

        # Add cluster information
        if self.config.section_enabled("cluster_information"):
            story.extend(
                self._create_cluster_information(
                    processed_data, page_tracker, "cluster_info" if is_first_pass else None
                )
            )
            story.append(PageBreak())

        # Add hardware inventory (includes Physical Rack Layout)
        if self.config.section_enabled("hardware_inventory"):
            hardware_section_key = "hardware_summary" if is_first_pass else None
            rack_layout_key = (
                "rack_layout"
                if (
                    is_first_pass
                    and processed_data.get("hardware_inventory", {}).get("rack_positions_available", False)
                )
                else None
            )
            story.extend(
                self._create_hardware_inventory(processed_data, page_tracker, hardware_section_key, rack_layout_key)
            )
            story.append(PageBreak())

        # Add comprehensive network configuration
        if self.config.section_enabled("network_configuration"):
            story.extend(
                self._create_comprehensive_network_configuration(
                    processed_data,
                    page_tracker,
                    "network_config" if is_first_pass else None,
                )
            )

        # Add switch configuration section
        if self.config.section_enabled("switch_configuration"):
            story.extend(
                self._create_switch_configuration(
                    processed_data, page_tracker, "switch_config" if is_first_pass else None
                )
            )

        # Add port mapping (if available and enabled in config)
        if self.config.section_enabled("port_mapping"):
            port_mapping_section = processed_data.get("sections", {}).get("port_mapping", {})
            port_mapping_data = port_mapping_section.get("data", {})
            port_mapping_enabled = (
                port_mapping_data.get("available") and port_mapping_section.get("status") == "complete"
            )

            if port_mapping_enabled:
                hardware = processed_data.get("hardware_inventory", {})
                switches = hardware.get("switches") or []
                story.extend(
                    self._create_port_mapping_section(
                        port_mapping_data,
                        switches,
                        page_tracker,
                        "port_mapping" if is_first_pass else None,
                    )
                )
                story.append(PageBreak())

                story.extend(
                    self._create_logical_network_diagram(
                        processed_data,
                        page_tracker,
                        "network_diagram" if is_first_pass else None,
                    )
                )
                story.append(PageBreak())
            else:
                story.append(PageBreak())
        else:
            story.append(PageBreak())

        # Add logical configuration
        if self.config.section_enabled("logical_configuration"):
            story.extend(
                self._create_logical_configuration(
                    processed_data,
                    page_tracker,
                    "logical_config" if is_first_pass else None,
                )
            )
            story.append(PageBreak())

        # Add security configuration
        if self.config.section_enabled("security_authentication"):
            story.extend(
                self._create_security_configuration(
                    processed_data,
                    page_tracker,
                    "security_config" if is_first_pass else None,
                )
            )
            story.append(PageBreak())

        # Health Check Results (optional)
        sections = processed_data.get("sections", {})
        health_data = sections.get("health_check", {}).get("data")
        if health_data and self.config.section_enabled("health_check"):
            story.extend(
                self._create_health_check_section(
                    health_data,
                    page_tracker,
                    "health_check" if is_first_pass else None,
                    processed_data=processed_data,
                )
            )
            story.append(PageBreak())

        # Post Deployment Activities (next steps checklist)
        activities_data = sections.get("post_deployment_activities", {}).get("data")
        if activities_data and self.config.section_enabled("post_deployment_activities"):
            story.extend(
                self._create_post_deployment_activities_section(
                    activities_data,
                    page_tracker,
                    "post_deploy_activities" if is_first_pass else None,
                    processed_data=processed_data,
                )
            )
            story.append(PageBreak())

        return story

    @staticmethod
    def _normalize_boxes_to_dict(boxes: Any) -> Dict[str, Any]:
        """Normalize box data to a dict keyed by name. Accepts list or dict."""
        if isinstance(boxes, dict):
            return boxes
        if isinstance(boxes, list):
            return {b.get("name") or str(b.get("id", i)): b for i, b in enumerate(boxes)}
        return {}

    def _ensure_hardware_inventory(self, data: Dict[str, Any]) -> None:
        """If hardware_inventory is missing or empty but raw 'hardware' exists, build it."""
        hi = data.get("hardware_inventory") or {}
        if not isinstance(hi, dict):
            return
        has_any = (
            (hi.get("cnodes") or [])
            or (hi.get("dnodes") or [])
            or (hi.get("cboxes") or {})
            or (hi.get("dboxes") or {})
            or (hi.get("switches") or [])
        )
        if has_any:
            return
        raw = data.get("hardware") or data.get("raw_hardware") or {}
        if not raw:
            return

        # Normalize to list when raw has dict-shaped cnodes/dnodes (e.g. id -> item)
        def _to_list(x: Any) -> List[Any]:
            if x is None:
                return []
            if isinstance(x, list):
                return x
            if isinstance(x, dict):
                return list(x.values())
            return []

        cnodes = _to_list(raw.get("cnodes"))
        dnodes = _to_list(raw.get("dnodes"))
        cboxes = self._normalize_boxes_to_dict(raw.get("cboxes") or {})
        dboxes = self._normalize_boxes_to_dict(raw.get("dboxes") or {})
        eboxes = self._normalize_boxes_to_dict(raw.get("eboxes") or {})
        switch_inv = data.get("switch_inventory") or data.get("raw_switch_inventory") or {}
        switches = (
            switch_inv.get("switches", [])
            if isinstance(switch_inv, dict)
            else (switch_inv if isinstance(switch_inv, list) else [])
        )
        data["hardware_inventory"] = {
            "cnodes": cnodes,
            "dnodes": dnodes,
            "cboxes": cboxes,
            "dboxes": dboxes,
            "eboxes": eboxes,
            "switches": switches,
            "total_nodes": len(cnodes) + len(dnodes),
            "rack_positions_available": bool(cnodes or dnodes or cboxes or dboxes),
            "physical_layout": None,
        }

    def _generate_with_reportlab(self, processed_data: Dict[str, Any], output_path: str) -> bool:
        """Generate PDF using ReportLab with two-pass generation for dynamic TOC."""
        try:
            self._ensure_hardware_inventory(processed_data)
            # Set up document with page template
            page_size = A4 if self.config.page_size == "A4" else letter

            # Create page template with footer
            generation_info = {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "completeness": processed_data.get("metadata", {}).get("overall_completeness", 0.0),
                "mgmt_vip": processed_data.get("cluster_summary", {}).get("mgmt_vip", "Unknown"),
            }
            page_template = self.brand_compliance.create_vast_page_template(
                generation_info,
                page_size=page_size,
                margins={
                    "margin_left": self.config.margin_left,
                    "margin_right": self.config.margin_right,
                    "margin_top": self.config.margin_top,
                    "margin_bottom": self.config.margin_bottom,
                },
                organization=getattr(self.config, "organization", None),
                include_page_numbers=getattr(self.config, "include_page_numbers", True),
            )

            em = self.brand_compliance.effective_margins
            self._frame_width = page_size[0] - em["left"] - em["right"]
            self._frame_height = page_size[1] - em["top"] - em["bottom"]

            # Create document with page template
            from reportlab.platypus import BaseDocTemplate
            import os
            import tempfile

            # PASS 1: Capture page numbers
            self.logger.info("First pass: Capturing page numbers for dynamic TOC...")
            page_tracker: Dict[str, int] = {}

            # Use mkstemp for first-pass temp PDF (Windows: NamedTemporaryFile can cause PermissionError on save)
            fd, temp_path = tempfile.mkstemp(suffix=".pdf")
            try:
                os.close(fd)
                temp_doc = BaseDocTemplate(
                    temp_path,
                    pagesize=page_size,
                    rightMargin=em["right"],
                    leftMargin=em["left"],
                    topMargin=em["top"],
                    bottomMargin=em["bottom"],
                )
                landscape_template = self._create_landscape_template(page_size, em)
                temp_doc.addPageTemplates([page_template, landscape_template])

                # Build story with page markers
                story_pass1 = self._build_report_story(processed_data, page_tracker)

                story_pass1.insert(0, NextPageTemplate("VastPage"))

                # Build temp PDF to capture page numbers
                temp_doc.build(story_pass1)
            finally:
                try:
                    os.unlink(temp_path)
                except OSError:
                    pass

            self.logger.info(f"First pass complete: Captured {len(page_tracker)} page numbers")
            # Log captured page numbers for verification
            for section_key, page_num in sorted(page_tracker.items()):
                self.logger.info(f"  {section_key}: page {page_num}")
            self.logger.debug(f"Page tracker: {page_tracker}")

            # PASS 2: Generate final PDF with dynamic TOC
            self.logger.info("Second pass: Generating final PDF with dynamic TOC...")

            doc = BaseDocTemplate(
                output_path,
                pagesize=page_size,
                rightMargin=em["right"],
                leftMargin=em["left"],
                topMargin=em["top"],
                bottomMargin=em["bottom"],
            )
            landscape_template = self._create_landscape_template(page_size, em)
            doc.addPageTemplates([page_template, landscape_template])

            # Build story with captured page numbers (no markers needed in second pass)
            story_pass2 = self._build_report_story(processed_data, page_tracker)

            story_pass2.insert(0, NextPageTemplate("VastPage"))

            # Build final PDF
            doc.build(story_pass2)

            self.logger.info(f"PDF report generated successfully: {output_path}")
            return True

        except Exception as e:
            self.logger.error(f"Error generating PDF with ReportLab: {e}")
            import traceback

            self.logger.error(traceback.format_exc())
            return False

    def _create_title_page(self, data: Dict[str, Any]) -> List[Any]:
        """Create VAST brand-compliant title page content."""
        content = []

        # Get cluster information
        cluster_info = data.get("cluster_summary", {})

        # Get hardware information from hardware inventory
        hardware_inventory = data.get("hardware_inventory", {})
        cnodes = hardware_inventory.get("cnodes") or []
        dnodes = hardware_inventory.get("dnodes") or []
        cboxes = hardware_inventory.get("cboxes") or {}
        dboxes = hardware_inventory.get("dboxes") or {}
        switches = hardware_inventory.get("switches") or []

        # Create VAST brand-compliant header
        title = "VAST As-Built Report"
        subtitle = "Customer Deployment Documentation"

        header_elements = self.brand_compliance.create_vast_header(
            title=title, subtitle=subtitle, cluster_info=cluster_info
        )
        content.extend(header_elements)

        # Add hardware information when any hardware is present (nodes, boxes, or switches)
        eboxes = hardware_inventory.get("eboxes") or {}
        if cnodes or dnodes or cboxes or dboxes or switches or eboxes:
            hardware_text = ""

            # EBox cluster: show EBox Hardware + EBox Quantity only (no CBox/DBox lines)
            if eboxes:
                ebox_vendors = set()
                for cnode in cnodes:
                    box_vendor = cnode.get("box_vendor", "") or ""
                    if box_vendor and box_vendor != "Unknown":
                        model = box_vendor.split(",")[0].strip()
                        if model:
                            ebox_vendors.add(model)
                if ebox_vendors:
                    hardware_text += f"<b>EBox Hardware:</b> {', '.join(sorted(ebox_vendors))}<br/>"
                hardware_text += f"<b>EBox Quantity:</b> {len(eboxes)}<br/>"
            else:
                # CBox Hardware: from CNode box_vendor when available, else from CBox data
                if cboxes:
                    cbox_vendors = set()
                    cbox_ids = set()
                    if cnodes:
                        for cnode in cnodes:
                            box_vendor = cnode.get("box_vendor", "Unknown")
                            if box_vendor and box_vendor != "Unknown":
                                cbox_vendors.add(box_vendor)
                            cbox_id = cnode.get("id")
                            if cbox_id:
                                cbox_ids.add(str(cbox_id))
                    if not cbox_vendors:
                        for _name, cbox_data in cboxes.items():
                            model = cbox_data.get("model") or cbox_data.get("hardware_type")
                            if model and model != "Unknown":
                                cbox_vendors.add(str(model))
                            cbox_id = cbox_data.get("id")
                            if cbox_id is not None:
                                cbox_ids.add(str(cbox_id))
                    if cbox_vendors:
                        hardware_text += f"<b>CBox Hardware:</b> {', '.join(sorted(cbox_vendors))}<br/>"
                    cbox_qty = len(cboxes) if cboxes else len(cbox_ids)
                    hardware_text += f"<b>CBox Quantity:</b> {cbox_qty}<br/>"

                # DBox Hardware (from DBox data using hardware_type)
                if dboxes:
                    dbox_hardware_types = set()
                    dbox_ids = set()
                    for dbox_name, dbox_data in dboxes.items():
                        hardware_type = dbox_data.get("hardware_type")
                        if hardware_type and hardware_type != "Unknown":
                            dbox_hardware_types.add(hardware_type)
                        dbox_id = dbox_data.get("id")
                        if dbox_id:
                            dbox_ids.add(str(dbox_id))
                    if dbox_hardware_types:
                        hardware_text += f"<b>DBox Hardware:</b> {', '.join(sorted(dbox_hardware_types))}<br/>"
                    hardware_text += f"<b>DBox Quantity:</b> {len(dbox_ids)}<br/>"

            # Switch Hardware (from switch inventory)
            switches = hardware_inventory.get("switches") or []
            if switches:
                switch_models = set()
                switch_count = len(switches)

                for switch in switches:
                    model = switch.get("model", "Unknown")
                    if model and model != "Unknown":
                        switch_models.add(model)

                if switch_models:
                    hardware_text += f"<b>Switch Hardware:</b> {', '.join(sorted(switch_models))}<br/>"

                hardware_text += f"<b>Switch Quantity:</b> {switch_count}"

            if hardware_text:
                # Create centered hardware information
                hardware_style = ParagraphStyle(
                    "CenteredHardware",
                    parent=self.brand_compliance.styles["vast_body"],
                    alignment=TA_CENTER,
                )
                hardware_para = Paragraph(hardware_text, hardware_style)
                content.append(hardware_para)
                content.append(Spacer(1, 20))

        # Remove the "Generated on" line as requested
        # Footer is now handled by page template

        return content

    def _create_table_of_contents_from_excel(self, excel_path: str) -> List[Any]:
        """Create table of contents from Excel file."""
        try:
            import openpyxl
            from reportlab.pdfbase.pdfmetrics import stringWidth
        except ImportError:
            self.logger.error("openpyxl not available for Excel import")
            return []

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "TOC_Title",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size + 2,
            spaceAfter=20,
            alignment=TA_LEFT,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            fontName=self._font("bold"),
        )

        content = []
        content.append(Paragraph("Table of Contents (from Excel)", title_style))
        content.append(Spacer(1, 8))

        # Load Excel file
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["Report-TOC"]
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {e}")
            content.append(Paragraph(f"Error loading Excel file: {e}", styles["Normal"]))
            return content

        # Read TOC data from Excel (A1:C60)
        available_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
        toc_table_data: list[Any] = []

        for row in range(1, 61):
            # Read values from columns A, B, C
            text_cell = ws[f"A{row}"]
            page_cell = ws[f"C{row}"]

            text = str(text_cell.value).strip() if text_cell.value else ""
            page_num = str(page_cell.value).strip() if page_cell.value else ""

            # Skip empty rows or header row
            if not text or text.lower() == "contents":
                continue

            # Determine indentation level by counting leading spaces
            # Support multiple indentation levels (each 4-5 spaces = 1 level)
            indent_level = 0
            original_text = str(text_cell.value) if text_cell.value else ""

            # Count leading spaces to determine indent level
            leading_spaces = len(original_text) - len(original_text.lstrip())
            if leading_spaces > 0:
                # Each 4-5 spaces = 1 indent level
                indent_level = (leading_spaces + 2) // 4  # Round to nearest level
                text = text.strip()

            # Read Excel cell formatting (bold, font size, color, alignment)
            # Default values
            is_bold_excel = False
            is_italic_excel = False
            text_size_excel = None
            text_color_excel = None
            text_alignment = TA_LEFT  # Default left alignment

            # Check if cell has font formatting
            if text_cell.font:
                is_bold_excel = text_cell.font.bold or False
                is_italic_excel = text_cell.font.italic or False

                # Get font size from Excel (convert from points)
                if text_cell.font.size:
                    text_size_excel = text_cell.font.size

                # Get font color from Excel
                if text_cell.font.color:
                    if text_cell.font.color.rgb:
                        # RGB format: "AARRGGBB" or "RRGGBB"
                        rgb = text_cell.font.color.rgb
                        if isinstance(rgb, str):
                            # Remove alpha channel if present
                            if len(rgb) == 8:
                                rgb = rgb[2:]  # Remove "AA" prefix
                            text_color_excel = colors.HexColor(f"#{rgb}")

            # Read cell alignment from Excel
            if text_cell.alignment:
                horizontal = text_cell.alignment.horizontal
                if horizontal == "center":
                    text_alignment = TA_CENTER
                elif horizontal == "right":
                    text_alignment = TA_RIGHT
                elif horizontal == "left":
                    text_alignment = TA_LEFT
                # Note: 'general' or None defaults to TA_LEFT

            # Determine if bold from Excel or fallback to page number logic
            is_bold = is_bold_excel if text_cell.font else bool(page_num)

            # Set font properties (use Excel formatting if available)
            if is_bold:
                text_font = self._font("bold")
                text_color = text_color_excel if text_color_excel else self.brand_compliance.colors.BACKGROUND_DARK
                text_size = text_size_excel if text_size_excel else (self.config.font_size - 1)
                page_font = self._font("bold")
                page_color = text_color if text_color else self.brand_compliance.colors.BACKGROUND_DARK
                page_size = text_size
                extra_space = 3
            else:
                text_font = self._font("italic") if is_italic_excel else self._font()
                text_color = text_color_excel if text_color_excel else colors.HexColor("#000000")
                text_size = text_size_excel if text_size_excel else (self.config.font_size - 2)
                page_font = text_font
                page_color = text_color
                page_size = text_size
                extra_space = 0

            # Check column B for special formatting (lines/separators)
            sep_cell = ws[f"B{row}"]
            separator_text = str(sep_cell.value).strip() if sep_cell.value else ""

            # Format text with indentation (only if not center/right aligned)
            if text_alignment == TA_LEFT:
                # Use 4 spaces per indent level for better visibility
                indent_space = "    " * indent_level if indent_level > 0 else ""
                full_text = f"{indent_space}{text}"
            else:
                full_text = text

            # Calculate dots if page number exists AND alignment is left
            # Center/right aligned entries don't get dots
            if page_num and text_alignment == TA_LEFT:
                text_width = stringWidth(full_text, text_font, text_size)
                page_width = stringWidth(page_num, page_font, page_size)

                spacing_buffer = 0.1 * inch
                available_for_dots = available_width - text_width - page_width - spacing_buffer

                dot_width = stringWidth(".", self._font(), text_size - 1)
                if dot_width > 0:
                    num_dots = int(available_for_dots / dot_width)
                    num_dots = max(3, num_dots)
                else:
                    num_dots = 30

                dots = '<font color="#CCCCCC">' + ("." * num_dots) + "</font>"
                text_with_dots = f"{full_text} {dots}"
            elif page_num and text_alignment == TA_CENTER:
                # Center-aligned with page number: text on left, number on right, centered line between
                if separator_text and separator_text in ["─", "-", "—", "_"]:
                    # Use separator character from column B
                    text_with_dots = f"{full_text} {separator_text * 80}"
                else:
                    # Default to dots but will be centered
                    text_with_dots = f"{full_text} {'.' * 80}"
            elif page_num and text_alignment == TA_RIGHT:
                # Right-aligned with page number
                text_with_dots = full_text
            else:
                text_with_dots = full_text

            # Create paragraph styles with Excel alignment
            text_style = ParagraphStyle(
                f"TOC_Text_{len(toc_table_data)}",
                parent=styles["Normal"],
                fontSize=text_size,
                fontName=text_font,
                textColor=text_color,
                alignment=text_alignment,  # Use Excel alignment
                spaceBefore=extra_space,
                spaceAfter=0.5,
                leading=text_size + 2,
            )

            # Page number alignment matches text alignment for center, otherwise right
            page_alignment = text_alignment if text_alignment == TA_CENTER else TA_RIGHT

            page_style = ParagraphStyle(
                f"TOC_Page_{len(toc_table_data)}",
                parent=styles["Normal"],
                fontSize=page_size,
                fontName=page_font,
                textColor=page_color,
                alignment=page_alignment,
                spaceBefore=extra_space,
                spaceAfter=0.5,
                leading=page_size + 2,
            )

            # Create paragraphs
            text_para = Paragraph(text_with_dots, text_style)
            if page_num:
                page_para = Paragraph(page_num, page_style)
            else:
                page_para = Paragraph("", page_style)

            toc_table_data.append([text_para, page_para])

        # Create table
        text_col_width = available_width - 0.5 * inch
        page_col_width = 0.5 * inch

        from reportlab.platypus import Table as RLTable

        toc_table = RLTable(toc_table_data, colWidths=[text_col_width, page_col_width])
        toc_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )

        content.append(toc_table)
        return content

    def _create_table_of_contents_dynamic(self, data: Dict[str, Any], page_tracker: Dict[str, int]) -> List[Any]:
        """
        Create table of contents with dynamically captured page numbers.

        Args:
            data: Processed cluster data
            page_tracker: Dictionary mapping section keys to page numbers

        Returns:
            List of flowables for the TOC
        """
        from reportlab.platypus import Table as RLTable
        from reportlab.pdfbase.pdfmetrics import stringWidth

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "TOC_Title",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size + 2,
            spaceAfter=20,
            alignment=TA_LEFT,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            fontName=self._font("bold"),
        )

        content = []
        content.append(Paragraph("Table of Contents", title_style))
        content.append(Spacer(1, 8))

        # TOC structure with section keys for dynamic page lookup
        # Format: (text, indent_level, section_key, is_bold)
        # section_key maps to page_tracker dictionary
        toc_structure = [
            # Executive Summary section
            ("Executive Summary", 0, "exec_summary", True),
            ("Cluster Overview", 1, None, False),
            ("Hardware Overview", 1, None, False),
            # Cluster Information section
            ("Cluster Information", 0, "cluster_info", True),
            ("Cluster Details", 1, None, False),
            ("Operational Status", 1, None, False),
            ("Feature Configuration", 1, None, False),
            # Hardware Summary section
            ("Hardware Summary", 0, "hardware_summary", True),
            ("Storage Capacity", 1, None, False),
            ("CBox Inventory", 1, None, False),
            ("DBox Inventory", 1, None, False),
            # Physical Rack Layout section (optional)
            ("Physical Rack Layout", 0, "rack_layout", True),
            # Network Configuration section
            ("Network Configuration", 0, "network_config", True),
            ("Cluster Network", 1, None, False),
            ("CNode Network", 1, None, False),
            ("DNode Network", 1, None, False),
            # Switch Configuration section
            ("Switch Configuration", 0, "switch_config", True),
            ("Switch Details", 1, None, False),
            ("Port Summary", 1, None, False),
            # Port Mapping section (optional)
            ("Port Mapping", 0, "port_mapping", True),
            ("Device Mapping", 1, None, False),
            # Logical Network Diagram section
            ("Logical Network Diagram", 0, "network_diagram", True),
            # Logical Configuration section
            ("Logical Configuration", 0, "logical_config", True),
            ("Tenants & Views", 1, None, False),
            ("Protection Policies", 1, None, False),
            # Security & Authentication section
            ("Security & Authentication", 0, "security_config", True),
            ("Encryption Configuration", 1, None, False),
            ("Authentication Services", 1, None, False),
            # Health Check section (optional)
            ("Cluster Health Check Results", 0, "health_check", True),
            # Post Deployment Activities section
            ("Post Deployment Activities", 0, "post_deploy_activities", True),
        ]

        # Map TOC section keys to config toggle keys so disabled sections
        # (and their child indent-1 items) are excluded from the TOC.
        _toc_key_to_config: Dict[str, str] = {
            "exec_summary": "executive_summary",
            "cluster_info": "cluster_information",
            "hardware_summary": "hardware_inventory",
            "rack_layout": "hardware_inventory",
            "network_config": "network_configuration",
            "switch_config": "switch_configuration",
            "port_mapping": "port_mapping",
            "network_diagram": "port_mapping",
            "logical_config": "logical_configuration",
            "security_config": "security_authentication",
            "health_check": "health_check",
            "post_deploy_activities": "post_deployment_activities",
        }

        # First pass: remove config-disabled sections and their children
        config_filtered: list[tuple[str, int, Optional[str], bool]] = []
        skip_children = False
        for entry in toc_structure:
            _text, indent_level, section_key, _is_bold = entry
            if indent_level == 0:
                cfg_key = _toc_key_to_config.get(section_key or "", "")
                skip_children = not self.config.section_enabled(cfg_key) if cfg_key else False
            if skip_children and indent_level > 0:
                continue
            if indent_level == 0 and skip_children:
                continue
            config_filtered.append(entry)

        # Second pass: remove optional sections that weren't captured in page_tracker
        filtered_structure: list[tuple[str, int, Optional[str], bool]] = []
        for entry in config_filtered:
            text, indent_level, section_key, is_bold = entry
            if section_key is None or section_key in page_tracker:
                filtered_structure.append(entry)

        # Build TOC table with calculated dot leaders
        available_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
        toc_table_data: list[Any] = []

        # List of subsections that should have extra space after them
        subsections_with_space_after = [
            "Hardware Overview",
            "Feature Configuration",
            "DBox Inventory",
            "Physical Rack Layout",
            "DNode Network",
            "Port Summary",
            "Device Mapping",
            "Logical Network Diagram",
            "Protection Policies",
            "Authentication Services",
            "Cluster Health Check Results",
        ]

        for idx, (text, indent_level, section_key, is_bold) in enumerate(filtered_structure):
            # Calculate indentation
            indent_space = "  " * indent_level if indent_level > 0 else ""
            full_text = f"{indent_space}{text}"

            # Get page number from tracker (if available)
            page_num = None
            if section_key and section_key in page_tracker:
                page_num = str(page_tracker[section_key])

            # Create styles for text and page number
            if is_bold:
                text_font = self._font("bold")
                text_color = self.brand_compliance.colors.BACKGROUND_DARK
                text_size = self.config.font_size - 1
                page_font = self._font("bold")
                page_color = self.brand_compliance.colors.BACKGROUND_DARK
                page_size = self.config.font_size - 1
                extra_space = 0 if idx == 0 else 12
            else:
                text_font = self._font()
                text_color = colors.HexColor("#000000")
                text_size = self.config.font_size - 2
                page_font = self._font()
                page_color = colors.HexColor("#000000")
                page_size = self.config.font_size - 2
                extra_space = 0

            # Add extra space after specific subsections
            if text in subsections_with_space_after:
                extra_space_after = 8.0
            else:
                extra_space_after = 0.5

            # Add dots and page numbers for entries that have page numbers
            if page_num:
                # Custom dot leader lengths for specific sections
                if text == "Executive Summary":
                    dot_leader_length = 5.0 * inch
                elif text == "Cluster Information":
                    dot_leader_length = 5.0 * inch
                elif text == "Hardware Summary":
                    dot_leader_length = 5.0 * inch
                elif text == "Physical Rack Layout":
                    dot_leader_length = 4.92 * inch
                elif text == "Network Configuration":
                    dot_leader_length = 4.87 * inch
                elif text == "Switch Configuration":
                    dot_leader_length = 4.95 * inch
                elif text == "Port Mapping":
                    dot_leader_length = 5.4 * inch
                elif text == "Logical Network Diagram":
                    dot_leader_length = 4.78 * inch
                elif text == "Logical Configuration":
                    dot_leader_length = 4.95 * inch
                else:
                    dot_leader_length = 4.75 * inch

                # Calculate how many dots fit in the specified space
                dot_width = stringWidth(".", self._font(), text_size - 1)
                if dot_width > 0:
                    num_dots = int(dot_leader_length / dot_width)
                    num_dots = max(3, num_dots)
                else:
                    num_dots = 150

                dots = '<font color="#CCCCCC">' + ("." * num_dots) + "</font>"
                text_with_dots = f"{full_text} {dots}"
            else:
                # Subsections without page numbers - no dots needed
                text_with_dots = full_text

            text_style = ParagraphStyle(
                f"TOC_Text_{len(toc_table_data)}",
                parent=styles["Normal"],
                fontSize=text_size,
                fontName=text_font,
                textColor=text_color,
                alignment=TA_LEFT,
                spaceBefore=extra_space,
                spaceAfter=extra_space_after,
                leading=text_size + 2,
            )

            page_style = ParagraphStyle(
                f"TOC_Page_{len(toc_table_data)}",
                parent=styles["Normal"],
                fontSize=page_size,
                fontName=page_font,
                textColor=page_color,
                alignment=TA_RIGHT,
                spaceBefore=extra_space,
                spaceAfter=extra_space_after,
                leading=page_size + 2,
            )

            # Create paragraphs
            text_para = Paragraph(text_with_dots, text_style)
            if page_num:
                page_para = Paragraph(page_num, page_style)
            else:
                page_para = Paragraph("", page_style)

            toc_table_data.append([text_para, page_para])

        # Create table
        text_col_width = available_width - 0.15 * inch
        page_col_width = 0.15 * inch

        toc_table = RLTable(toc_table_data, colWidths=[text_col_width, page_col_width])
        toc_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )

        content.append(toc_table)
        return content

    def _create_table_of_contents(self, data: Dict[str, Any]) -> List[Any]:
        """Create enhanced table of contents with dot leaders and perfect alignment."""
        from reportlab.platypus import Table as RLTable

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "TOC_Title",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size + 2,
            spaceAfter=20,
            alignment=TA_LEFT,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            fontName=self._font("bold"),
        )

        content = []

        content.append(Paragraph("Table of Contents", title_style))
        content.append(Spacer(1, 8))

        # TOC structure with sections, subsections, and page numbers
        # Format: (text, indent_level, page_num, is_bold)
        # Structure: Parent section first, then its subsections indented below
        toc_structure = [
            # Executive Summary section
            ("Executive Summary", 0, "3", True),
            ("Cluster Overview", 1, None, False),
            ("Hardware Overview", 1, None, False),
            # Cluster Information section
            ("Cluster Information", 0, "4", True),
            ("Cluster Details", 1, None, False),
            ("Operational Status", 1, None, False),
            ("Feature Configuration", 1, None, False),
            # Hardware Summary section
            ("Hardware Summary", 0, "5", True),
            ("Storage Capacity", 1, None, False),
            ("CBox Inventory", 1, None, False),
            ("DBox Inventory", 1, None, False),
            # Physical Rack Layout section (no subsections)
            ("Physical Rack Layout", 0, "6", True),
            # Network Configuration section
            ("Network Configuration", 0, "7", True),
            ("Cluster Network", 1, None, False),
            ("CNode Network", 1, None, False),
            ("DNode Network", 1, None, False),
            # Switch Configuration section
            ("Switch Configuration", 0, "8", True),
            ("Switch Details", 1, None, False),
            ("Port Summary", 1, None, False),
            # Port Mapping section
            ("Port Mapping", 0, "10", True),
            ("Device Mapping", 1, None, False),
            # Logical Network Diagram section (no subsections)
            ("Logical Network Diagram", 0, "11", True),
            # Logical Configuration section
            ("Logical Configuration", 0, "12", True),
            ("Tenants & Views", 1, None, False),
            ("Protection Policies", 1, None, False),
            # Security & Authentication section
            ("Security & Authentication", 0, "13", True),
            ("Encryption Configuration", 1, None, False),
            ("Authentication Services", 1, None, False),
        ]

        _label_to_config: Dict[str, str] = {
            "Executive Summary": "executive_summary",
            "Cluster Information": "cluster_information",
            "Hardware Summary": "hardware_inventory",
            "Physical Rack Layout": "hardware_inventory",
            "Network Configuration": "network_configuration",
            "Switch Configuration": "switch_configuration",
            "Port Mapping": "port_mapping",
            "Logical Network Diagram": "port_mapping",
            "Logical Configuration": "logical_configuration",
            "Security & Authentication": "security_authentication",
        }
        config_filtered_ph: list[tuple[str, int, Optional[str], bool]] = []
        skip_children = False
        for entry in toc_structure:
            _text, indent_level, _page, _bold = entry
            if indent_level == 0:
                cfg_key = _label_to_config.get(_text, "")
                skip_children = not self.config.section_enabled(cfg_key) if cfg_key else False
            if skip_children:
                continue
            config_filtered_ph.append(entry)
        toc_structure = config_filtered_ph

        # Build TOC table with calculated dot leaders for perfect alignment
        from reportlab.pdfbase.pdfmetrics import stringWidth

        available_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
        toc_table_data: list[Any] = []

        # List of subsections that should have extra space after them (to separate section groups)
        subsections_with_space_after = [
            "Hardware Overview",
            "Feature Configuration",
            "DBox Inventory",
            "Physical Rack Layout",
            "DNode Network",
            "Port Summary",
            "Device Mapping",
            "Logical Network Diagram",
            "Protection Policies",
            "Authentication Services",
            "Cluster Health Check Results",
        ]

        for idx, (text, indent_level, page_num, is_bold) in enumerate(toc_structure):
            # Calculate indentation using spaces (smaller indent for compact view)
            indent_space = "  " * indent_level if indent_level > 0 else ""
            full_text = f"{indent_space}{text}"

            # Create styles for text and page number
            if is_bold:
                text_font = self._font("bold")
                text_color = self.brand_compliance.colors.BACKGROUND_DARK
                text_size = self.config.font_size - 1  # Slightly smaller for compact view
                page_font = self._font("bold")
                page_color = self.brand_compliance.colors.BACKGROUND_DARK
                page_size = self.config.font_size - 1
                # More space before main sections (except first one) to separate from subsections above
                extra_space = 0 if idx == 0 else 12
            else:
                text_font = self._font()
                text_color = colors.HexColor("#000000")
                text_size = self.config.font_size - 2  # Smaller for subsections
                page_font = self._font()
                page_color = colors.HexColor("#000000")
                page_size = self.config.font_size - 2
                extra_space = 0

            # Add extra space after specific subsections to separate section groups
            if text in subsections_with_space_after:
                extra_space_after = 8.0  # Extra space after these subsections
            else:
                extra_space_after = 0.5

            # Only add dots and page numbers for entries that have page numbers
            if page_num:
                # Custom dot leader lengths for specific sections
                if text == "Executive Summary":
                    dot_leader_length = 5.0 * inch
                elif text == "Cluster Information":
                    dot_leader_length = 5.0 * inch
                elif text == "Hardware Summary":
                    dot_leader_length = 5.0 * inch
                elif text == "Physical Rack Layout":
                    dot_leader_length = 4.92 * inch
                elif text == "Network Configuration":
                    dot_leader_length = 4.87 * inch
                elif text == "Switch Configuration":
                    dot_leader_length = 4.95 * inch
                elif text == "Port Mapping":
                    dot_leader_length = 5.4 * inch
                elif text == "Logical Network Diagram":
                    dot_leader_length = 4.78 * inch
                elif text == "Logical Configuration":
                    dot_leader_length = 4.95 * inch
                else:
                    dot_leader_length = 4.75 * inch  # Default for all other entries

                # Calculate how many dots fit in the specified space
                dot_width = stringWidth(".", self._font(), text_size - 1)
                if dot_width > 0:
                    num_dots = int(dot_leader_length / dot_width)
                    num_dots = max(3, num_dots)  # Minimum 3 dots
                else:
                    num_dots = 150  # Fallback for 5 inches

                dots = '<font color="#CCCCCC">' + ("." * num_dots) + "</font>"
                text_with_dots = f"{full_text} {dots}"
            else:
                # Subsections without page numbers - no dots needed
                text_with_dots = full_text

            text_style = ParagraphStyle(
                f"TOC_Text_{len(toc_table_data)}",
                parent=styles["Normal"],
                fontSize=text_size,
                fontName=text_font,
                textColor=text_color,
                alignment=TA_LEFT,
                spaceBefore=extra_space,
                spaceAfter=extra_space_after,  # Variable spacing based on subsection
                leading=text_size + 2,  # Compact line spacing
            )

            page_style = ParagraphStyle(
                f"TOC_Page_{len(toc_table_data)}",
                parent=styles["Normal"],
                fontSize=page_size,
                fontName=page_font,
                textColor=page_color,
                alignment=TA_RIGHT,
                spaceBefore=extra_space,
                spaceAfter=extra_space_after,  # Variable spacing based on subsection
                leading=page_size + 2,
            )

            # Create paragraphs
            text_para = Paragraph(text_with_dots, text_style)
            if page_num:
                page_para = Paragraph(page_num, page_style)
            else:
                page_para = Paragraph("", page_style)  # Empty for subsections

            toc_table_data.append([text_para, page_para])

        # Create a 2-column table: text+dots | page number
        # The text column should be wide, page column narrow
        # Page numbers positioned close to dot leaders (0.15" from right edge)
        text_col_width = available_width - 0.15 * inch
        page_col_width = 0.15 * inch

        toc_table = RLTable(toc_table_data, colWidths=[text_col_width, page_col_width])
        toc_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),  # Text column left-aligned
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),  # Page column right-aligned
                    ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),  # Align to baseline
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )

        content.append(toc_table)

        return content

    def _create_executive_summary(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """Create VAST brand-compliant executive summary section."""
        content = []

        # Add section heading with VAST styling
        heading_elements = self.brand_compliance.create_vast_section_heading("Executive Summary", level=1)
        content.extend(heading_elements)

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        styles = getSampleStyleSheet()
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "This VAST As-Built Report provides a comprehensive technical documentation of the deployed VAST Data cluster infrastructure, configuration, and operational status. The report serves as a critical reference for system administrators, storage engineers, and technical stakeholders to understand the current state of the cluster deployment, validate configuration compliance, and support ongoing operations and troubleshooting. The Executive Summary consolidates key operational metrics, hardware inventory, and cluster health indicators into high-level overview tables that enable rapid assessment of cluster status and capacity utilization.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        # Cluster overview table - resequenced to match screenshot order
        cluster_info = data.get("cluster_summary", {})
        cluster_id = cluster_info.get("cluster_id", "Unknown")
        cluster_name = cluster_info.get("name", "Unknown")
        mgmt_vip = cluster_info.get("mgmt_vip", "Unknown")
        url = cluster_info.get("url", "Unknown")
        build = cluster_info.get("build", "Unknown")
        psnt = cluster_info.get("psnt", "Unknown")
        guid = cluster_info.get("guid", "Unknown")
        uptime = cluster_info.get("uptime", "Unknown")
        online_start_time = cluster_info.get("online_start_time", "Unknown")
        deployment_time = cluster_info.get("deployment_time", "Unknown")

        cluster_overview_data = [
            ["ID", cluster_id],
            ["Name", cluster_name],
            ["Management VIP", mgmt_vip],
            ["URL", url],
            ["Build", build],
            ["PSNT", psnt],
            ["GUID", guid],
            ["Uptime", uptime],
            ["Online Since", online_start_time],
            ["Deployed", deployment_time],
        ]

        # Create cluster overview table with same style as Cluster Information
        cluster_table_elements = self._create_cluster_info_table(cluster_overview_data, "Cluster Overview")
        content.extend(cluster_table_elements)
        content.append(Spacer(1, 12))

        # Hardware overview table
        hardware = data.get("hardware_inventory", {})
        total_nodes = hardware.get("total_nodes", 0)
        cnodes = len(hardware.get("cnodes") or [])
        dnodes = len(hardware.get("dnodes") or [])
        cboxes_raw = hardware.get("cboxes") or {}
        dboxes_raw = hardware.get("dboxes") or {}
        eboxes_raw = hardware.get("eboxes") or {}
        cboxes_count = len(cboxes_raw)
        dboxes_count = len(dboxes_raw)
        eboxes = len(eboxes_raw)
        switches_list = hardware.get("switches") or []
        total_switches = len(switches_list)

        # EBox clusters: when any EBoxes are present, show CBoxes=0 and DBoxes=0 in Overview (per EBOX-HARDWARE-TABLE-IMPLEMENTATION-PLAN)
        is_ebox_cluster = bool(eboxes_raw)
        if is_ebox_cluster:
            cboxes_display = 0
            dboxes_display = 0
        else:
            cboxes_display = cboxes_count
            dboxes_display = dboxes_count

        # Calculate leaf and spine switches
        # Logic: If 2 switches = 2 leaf, if 4 = 2 leaf + 2 spine, if >4 = 2 spine + rest are leaf
        if total_switches == 2:
            leaf_switches = 2
            spine_switches = 0
        elif total_switches == 4:
            leaf_switches = 2
            spine_switches = 2
        elif total_switches > 4:
            spine_switches = 2
            leaf_switches = total_switches - 2
        else:
            leaf_switches = total_switches
            spine_switches = 0

        hardware_overview_data = [
            ["CBoxes", str(cboxes_display)],
            ["CNodes", str(cnodes)],
            ["", ""],  # Empty line
            ["DBoxes", str(dboxes_display)],
            ["DNodes", str(dnodes)],
            ["", ""],  # Empty line
            ["EBoxes", str(eboxes)],
            ["", ""],  # Empty line
            ["Switches", str(total_switches)],
            ["Leaf", str(leaf_switches)],
            ["Spine", str(spine_switches)],
        ]

        # Create hardware overview table with same style as Cluster Information
        hardware_table_elements = self._create_cluster_info_table(hardware_overview_data, "Hardware Overview")
        content.extend(hardware_table_elements)
        content.append(Spacer(1, 12))

        return content

    def _create_cluster_info_table(self, table_data: List[List[str]], title: str) -> List[Any]:
        """Create a cluster information style table with VAST branding."""
        content = []

        # Create table with VAST styling
        table_elements = self.brand_compliance.create_vast_table(table_data, title, ["Description", "Value"])
        content.extend(table_elements)

        return content

    def _create_ebox_only_inventory_table(
        self,
        eboxes: Dict[str, Any],
        cnodes: List[Dict[str, Any]],
        dnodes: List[Dict[str, Any]],
        switches: List[Dict[str, Any]],
    ) -> Optional[List[Any]]:
        """
        Build Hardware Inventory table for EBox-only clusters: 5 columns (no Node),
        row order EBox → CNode → 2× DNodes per EBox, model from CNode box_vendor, switches at bottom.
        Returns None if no rows (e.g. no eboxes with rack_unit).
        """
        headers_5 = ["Rack", "Model", "Name/Serial Number", "Status", "Height"]
        ebox_list = list(eboxes.values())

        def ebox_sort_key(e: Dict[str, Any]) -> tuple:
            rack = e.get("rack_name") or "Unknown"
            ru = e.get("rack_unit") or ""
            try:
                u_val = int(ru.upper().replace("U", ""))
            except (TypeError, ValueError):
                u_val = 0
            return (rack, -u_val)

        ebox_list.sort(key=ebox_sort_key)
        all_rows: List[List[str]] = []

        for ebox in ebox_list:
            ebox_id = ebox.get("id")
            rack_name = ebox.get("rack_name") or "Unknown"
            ebox_name = ebox.get("name", "Unknown")
            ebox_state = ebox.get("state", "Unknown")
            rack_unit = ebox.get("rack_unit") or "1U"

            # Find associated CNode and DNodes by ebox_id
            cnode = None
            for c in cnodes:
                if c.get("ebox_id") == ebox_id:
                    cnode = c
                    break
            model = "Unknown"
            if cnode:
                bv = cnode.get("box_vendor") or "Unknown"
                model = (bv.split(",")[0].strip()) if isinstance(bv, str) and bv else "Unknown"

            # EBox row
            all_rows.append([rack_name, model, ebox_name, ebox_state, rack_unit])

            # CNode row (no Model value)
            if cnode:
                cnode_name = cnode.get("name") or f"cnode-{cnode.get('id', 'Unknown')}"
                cnode_status = cnode.get("status", "Unknown")
                all_rows.append([rack_name, "", cnode_name, cnode_status, rack_unit])
            else:
                all_rows.append([rack_name, "", "N/A", "Unknown", rack_unit])

            # Up to 2 DNode rows (no Model value; Status = ACTIVE/FAILED from API)
            matching_dnodes = sorted(
                [d for d in dnodes if d.get("ebox_id") == ebox_id],
                key=lambda d: d.get("name") or "",
                reverse=True,
            )
            for d in matching_dnodes[:2]:
                dnode_name = d.get("name") or f"dnode-{d.get('id', 'Unknown')}"
                dnode_state = d.get("status") or d.get("state") or ""
                all_rows.append([rack_name, "", dnode_name, dnode_state, rack_unit])
            if len(matching_dnodes) == 0:
                all_rows.append([rack_name, "", "N/A", "", rack_unit])

        # Switches at bottom
        if switches:
            manual_rack_map = getattr(self, "manual_rack_placements", {})
            manual_sw_rack: Dict[str, str] = {}
            if manual_rack_map:
                for rn, placements in manual_rack_map.items():
                    for p in placements:
                        sw_name = p.get("name", p.get("hostname", ""))
                        if sw_name:
                            manual_sw_rack[sw_name] = rn
            default_rack = "Unknown"
            if not manual_sw_rack:
                if hasattr(self, "switch_rack_name") and self.switch_rack_name:
                    default_rack = self.switch_rack_name
                elif ebox_list:
                    default_rack = ebox_list[0].get("rack_name") or "Unknown"
            for switch_num, switch in enumerate(switches, start=1):
                sw_name = switch.get("name", switch.get("hostname", "Unknown"))
                hostname = switch.get("hostname", sw_name)
                model_sw = switch.get("model", "Unknown")
                if isinstance(model_sw, str) and "," in model_sw:
                    model_sw = model_sw.split(",")[0].strip()
                serial = switch.get("serial", "Unknown")
                state_sw = switch.get("state", "Unknown")
                rack_name_sw = manual_sw_rack.get(sw_name, default_rack)
                position = ""
                if hasattr(self, "switch_positions") and switch_num in self.switch_positions:
                    position = f"U{self.switch_positions[switch_num]}"
                all_rows.append([rack_name_sw, model_sw, serial, state_sw, position])

        if not all_rows:
            return None
        return cast(
            List[Any],
            self.brand_compliance.create_vast_hardware_table_with_pagination(all_rows, "Hardware Inventory", headers_5),
        )

    def _create_consolidated_inventory_table(
        self,
        cboxes: Dict[str, Any],
        cnodes: List[Dict[str, Any]],
        dboxes: Dict[str, Any],
        dnodes: List[Dict[str, Any]],
        switches: List[Dict[str, Any]],
        eboxes: Optional[Dict[str, Any]] = None,
    ) -> List[Any]:
        """
        Create consolidated hardware inventory table with CBoxes, DBoxes, EBoxes, and Switches.
        Hardware is grouped by rack name. Creates one row per CNode/DNode.

        Args:
            cboxes: CBox data from /api/v7/cboxes/
            cnodes: CNode data from /api/v7/cnodes/
            dboxes: DBox data from /api/v7/dboxes/
            dnodes: DNode data from /api/v7/dnodes/
            switches: Switch data from switch inventory
            eboxes: EBox data from /api/v7/eboxes/ (optional)

        Returns:
            List[Any]: Table elements
        """
        eboxes = eboxes or {}
        # EBox-only path: when cluster has EBoxes, use 5-column table (no Node column),
        # row order EBox → CNode → 2× DNodes per EBox, model from CNode vendor, switches at bottom.
        is_ebox_cluster = bool(eboxes)
        if is_ebox_cluster:
            ebox_only_elements = self._create_ebox_only_inventory_table(eboxes, cnodes, dnodes, switches)
            if ebox_only_elements is not None:
                return ebox_only_elements
            # Fall through to standard table if no rows produced (e.g. no rack_unit)

        # Prepare table data - will be grouped by rack
        all_rows = []
        headers = ["Rack", "Node", "Model", "Name/Serial Number", "Status", "Height"]

        # Build mapping of cbox_id to rack_name for CNodes
        cbox_id_to_rack_name = {}
        for cbox_name, cbox_data in cboxes.items():
            cbox_id = cbox_data.get("id")
            rack_name = cbox_data.get("rack_name") or "Unknown"
            if cbox_id:
                cbox_id_to_rack_name[cbox_id] = rack_name

        # Build mapping of dbox_id to rack_name for DNodes
        dbox_id_to_rack_name = {}
        for dbox_name, dbox_data in dboxes.items():
            dbox_id = dbox_data.get("id")
            rack_name = dbox_data.get("rack_name") or "Unknown"
            if dbox_id:
                dbox_id_to_rack_name[dbox_id] = rack_name

        # Add CBoxes
        if cboxes and cnodes:
            # Create a mapping of cbox_id to box_vendor and status from cnodes
            cbox_vendor_map = {}
            cbox_status_map = {}
            # Create a mapping of cbox_id to list of cnode names
            cbox_to_cnode_names: dict[str, list[str]] = {}
            # CNode name -> serial (Dell Asset Tag) for dell_turin_cbox Model column
            cnode_name_to_serial = {}
            for cnode in cnodes:
                cbox_id = cnode.get("cbox_id")
                box_vendor = cnode.get("box_vendor", "Unknown")
                status = cnode.get("status", "Unknown")
                # Use name field (programmatically generated) not hostname (customer-assigned)
                # Name is based on deployment index (e.g., cnode-3-10, cnode-3-11, cnode-3-12)
                cnode_name = cnode.get("name") or f"cnode-{cnode.get('id', 'Unknown')}"
                serial = cnode.get("serial_number") or cnode.get("serial") or ""
                if serial:
                    cnode_name_to_serial[cnode_name] = serial
                if cbox_id:
                    cbox_vendor_map[cbox_id] = box_vendor
                    cbox_status_map[cbox_id] = status
                    # Build list of CNode names for this CBox
                    if cbox_id not in cbox_to_cnode_names:
                        cbox_to_cnode_names[cbox_id] = []
                    cbox_to_cnode_names[cbox_id].append(cnode_name)

            cbox_rows = []
            for cbox_name, cbox_data in cboxes.items():
                cbox_id = cbox_data.get("id", "Unknown")
                name = cbox_data.get("name", "Unknown")
                rack_unit = cbox_data.get("rack_unit", "Unknown")
                rack_name = cbox_data.get("rack_name") or "Unknown"

                # Get model and status from cnodes data using cbox_id (strip NIC description after comma)
                model = cbox_vendor_map.get(cbox_id, "Unknown")
                if isinstance(model, str) and "," in model:
                    model = model.split(",")[0].strip()
                status = cbox_status_map.get(cbox_id, "Unknown")

                # Get CNode names for this CBox - create one row per CNode with Model
                cnode_names = cbox_to_cnode_names.get(cbox_id, [])
                if cnode_names:
                    # Create one row for each CNode
                    for cnode_name in cnode_names:
                        row = [rack_name, cnode_name, model, name, status, rack_unit]
                        cbox_rows.append((rack_name, cbox_id, cnode_name, row))
                else:
                    # No CNodes found, create one row with N/A
                    row = [rack_name, "N/A", model, name, status, rack_unit]
                    cbox_rows.append((rack_name, cbox_id, "N/A", row))

            # Sort by rack name, then by numeric ID, then by CNode name
            cbox_rows.sort(key=lambda x: (x[0], int(x[1]) if str(x[1]).isdigit() else 0, x[2]))
            all_rows.extend([row for _, _, _, row in cbox_rows])

        # Add DBoxes with DNode names
        if dboxes and dnodes:
            # Create a mapping of dbox_id to list of (dnode_name, dnode_state)
            dbox_to_dnodes: dict[str, list[tuple[str, str]]] = {}
            for dnode in dnodes:
                dbox_id = dnode.get("dbox_id")
                dnode_name = dnode.get("name") or f"dnode-{dnode.get('id', 'Unknown')}"
                dnode_state = dnode.get("status") or dnode.get("state") or "Unknown"
                if dbox_id:
                    if dbox_id not in dbox_to_dnodes:
                        dbox_to_dnodes[dbox_id] = []
                    dbox_to_dnodes[dbox_id].append((dnode_name, dnode_state))

            dbox_rows = []
            for dbox_name, dbox_data in dboxes.items():
                dbox_id = dbox_data.get("id", "Unknown")
                hardware_type = dbox_data.get("hardware_type", "Unknown")
                if isinstance(hardware_type, str) and "," in hardware_type:
                    hardware_type = hardware_type.split(",")[0].strip()
                name = dbox_data.get("name", "Unknown")
                state = dbox_data.get("state", "Unknown")
                rack_unit = dbox_data.get("rack_unit", "Unknown")
                rack_name = dbox_data.get("rack_name") or "Unknown"

                # Get DNodes for this DBox - create one row per DNode with Model and Status (ACTIVE/FAILED)
                dnode_list = dbox_to_dnodes.get(dbox_id, [])
                if dnode_list:
                    for dnode_name, dnode_state in dnode_list:
                        row = [
                            rack_name,
                            dnode_name,
                            hardware_type,
                            name,
                            dnode_state,
                            rack_unit,
                        ]
                        dbox_rows.append((rack_name, dbox_id, dnode_name, row))
                else:
                    # No DNodes found, create one row with N/A
                    row = [rack_name, "N/A", hardware_type, name, state, rack_unit]
                    dbox_rows.append((rack_name, dbox_id, "N/A", row))

            # Sort by rack name, then by numeric ID, then by DNode name
            dbox_rows.sort(key=lambda x: (x[0], int(x[1]) if str(x[1]).isdigit() else 0, x[2]))
            all_rows.extend([row for _, _, _, row in dbox_rows])
        elif dboxes:
            # Fallback if no dnodes data available
            dbox_rows = []
            for dbox_name, dbox_data in dboxes.items():
                dbox_id = dbox_data.get("id", "Unknown")
                hardware_type = dbox_data.get("hardware_type", "Unknown")
                if isinstance(hardware_type, str) and "," in hardware_type:
                    hardware_type = hardware_type.split(",")[0].strip()
                name = dbox_data.get("name", "Unknown")
                state = dbox_data.get("state", "Unknown")
                rack_unit = dbox_data.get("rack_unit", "Unknown")
                rack_name = dbox_data.get("rack_name") or "Unknown"

                # Create row data with N/A for DNode column (4-tuple to match dnode branch)
                row = [rack_name, "N/A", hardware_type, name, state, rack_unit]
                dbox_rows.append((rack_name, dbox_id, "N/A", row))

            # Sort by rack name, then by numeric ID
            dbox_rows.sort(key=lambda x: (x[0], int(x[1]) if str(x[1]).isdigit() else 0))
            all_rows.extend([row for _, _, _, row in dbox_rows])

        # Add EBoxes (enclosures; one row per EBox)
        if eboxes:
            ebox_rows = []
            for ebox_name, ebox_data in eboxes.items():
                rack_name = ebox_data.get("rack_name") or "Unknown"
                name = ebox_data.get("name", "Unknown")
                state = ebox_data.get("state", "Unknown")
                rack_unit = ebox_data.get("rack_unit", "N/A")
                row = [rack_name, "EBox", "Enclosure", name, state, rack_unit]
                ebox_rows.append((rack_name, ebox_data.get("id"), row))
            ebox_rows.sort(key=lambda x: (x[0], int(x[1]) if x[1] is not None and str(x[1]).isdigit() else 0))
            all_rows.extend([row for _, _, row in ebox_rows])

        # Add Switches
        if switches:
            manual_rack_map = getattr(self, "manual_rack_placements", {})

            # Build per-switch rack lookup from manual placements
            manual_sw_rack: Dict[str, str] = {}
            if manual_rack_map:
                for rn, placements in manual_rack_map.items():
                    for p in placements:
                        sw_name = p.get("name", p.get("hostname", ""))
                        if sw_name:
                            manual_sw_rack[sw_name] = rn

            if not manual_sw_rack:
                if hasattr(self, "switch_rack_name") and self.switch_rack_name:
                    default_rack = self.switch_rack_name
                else:
                    known_rack_names = sorted(
                        set(list(cbox_id_to_rack_name.values()) + list(dbox_id_to_rack_name.values()))
                    )
                    default_rack = known_rack_names[0] if known_rack_names else "Unknown"
            else:
                default_rack = "Unknown"

            switch_rows = []
            for switch_num, switch in enumerate(switches, start=1):
                sw_name = switch.get("name", switch.get("hostname", "Unknown"))
                hostname = switch.get("hostname", sw_name)
                model = switch.get("model", "Unknown")
                if isinstance(model, str) and "," in model:
                    model = model.split(",")[0].strip()
                serial = switch.get("serial", "Unknown")
                state = switch.get("state", "Unknown")

                rack_name = manual_sw_rack.get(sw_name, default_rack)

                position = ""
                if hasattr(self, "switch_positions") and switch_num in self.switch_positions:
                    u_pos = self.switch_positions[switch_num]
                    position = f"U{u_pos}"

                row = [rack_name, "N/A", model, serial, state, position]
                switch_rows.append((rack_name, hostname, row))

            switch_rows.sort(key=lambda x: (x[0], x[1]))

            for idx, (_rn, _hn, row) in enumerate(switch_rows, start=1):
                if hasattr(self, "switch_positions") and idx in self.switch_positions:
                    u_pos = self.switch_positions[idx]
                    row[5] = f"U{u_pos}"

            all_rows.extend([row for _, _, row in switch_rows])

        if not all_rows:
            return []

        def _parse_u(val: str) -> int:
            """Parse 'U24' → 24, return 0 on failure."""
            try:
                return int(str(val).upper().replace("U", ""))
            except (TypeError, ValueError):
                return 0

        def _dev_order(row) -> int:
            """Device-type priority: CBox/EBox=0, CNode=1, DNode=2, DBox(no-node)=3, Switch=4."""
            node_col = row[1] or ""
            name_col = row[3] or ""
            if node_col == "EBox":
                return 0
            if node_col.startswith("cnode-"):
                return 1
            if node_col.startswith("dnode-"):
                return 2
            if name_col.startswith("cbox-") or name_col.startswith("CB-"):
                return 0
            if name_col.startswith("dbox-") or "DB-" in name_col:
                return 0
            if node_col == "N/A":
                return 3
            return 4

        # Two-pass stable sort: name descending first, then primary criteria ascending.
        # Python's stable sort preserves name order within identical primary keys.
        all_rows.sort(key=lambda r: (r[3] or r[1] or ""), reverse=True)
        all_rows.sort(key=lambda r: (r[0] or "Unknown", -_parse_u(r[5] or ""), _dev_order(r)))

        # Create table with VAST styling
        return cast(
            List[Any],
            self.brand_compliance.create_vast_hardware_table_with_pagination(all_rows, "Hardware Inventory", headers),
        )

    def _create_cluster_information(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """Create VAST brand-compliant cluster information section."""
        content = []

        # Add section heading with VAST styling
        heading_elements = self.brand_compliance.create_vast_section_heading("Cluster Information", level=1)
        content.extend(heading_elements)

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        styles = getSampleStyleSheet()
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Cluster Information section provides detailed operational status and configuration parameters for the VAST Data cluster. This section captures essential cluster metadata including cluster identification, operational state, management network configuration, and feature flags that define the cluster's capabilities and current operational mode. The information presented here is critical for understanding the cluster's current operational status, validating proper configuration, and supporting troubleshooting activities. This data is collected directly from the cluster's management API and represents the real-time operational state of the system.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        cluster_info = data.get("cluster_summary", {})

        # Create cluster info table with VAST styling - reorganized per requirements
        cluster_name = cluster_info.get("name", "Unknown")

        # Get capacity display format for this table
        capacity_base_10 = cluster_info.get("capacity_base_10", None)
        if capacity_base_10 is True:
            capacity_format = "True"
        elif capacity_base_10 is False:
            capacity_format = "False"
        else:
            capacity_format = "Unknown"

        cluster_data = [
            ["State", cluster_info.get("state", "Unknown")],
            ["SSD RAID State", cluster_info.get("ssd_raid_state", "Unknown")],
            ["NVRAM RAID State", cluster_info.get("nvram_raid_state", "Unknown")],
            ["Memory RAID State", cluster_info.get("memory_raid_state", "Unknown")],
            ["Leader State", cluster_info.get("leader_state", "Unknown")],
            ["Leader CNode", cluster_info.get("leader_cnode", "Unknown")],
            ["Management CNode", cluster_info.get("mgmt_cnode", "Unknown")],
            ["Management Inner VIP", cluster_info.get("mgmt_inner_vip", "Unknown")],
            [
                "Management Inner VIP CNode",
                cluster_info.get("mgmt_inner_vip_cnode", "Unknown"),
            ],
            [
                "Enabled",
                (
                    "Yes"
                    if cluster_info.get("enabled")
                    else ("No" if cluster_info.get("enabled") is not None else "Unknown")
                ),
            ],
            [
                "Similarity Enabled",
                (
                    "Yes"
                    if cluster_info.get("enable_similarity")
                    else ("No" if cluster_info.get("enable_similarity") is not None else "Unknown")
                ),
            ],
            [
                "Write-Back RAID Enabled",
                (
                    "Yes"
                    if cluster_info.get("is_wb_raid_enabled")
                    else ("No" if cluster_info.get("is_wb_raid_enabled") is not None else "Unknown")
                ),
            ],
            [
                "Write-Back RAID Layout",
                cluster_info.get("wb_raid_layout", "Unknown"),
            ],
            [
                "DBox HA Support",
                (
                    "Yes"
                    if cluster_info.get("dbox_ha_support")
                    else ("No" if cluster_info.get("dbox_ha_support") is not None else "Unknown")
                ),
            ],
            [
                "Rack Level Resiliency",
                (
                    "Yes"
                    if cluster_info.get("enable_rack_level_resiliency")
                    else ("No" if cluster_info.get("enable_rack_level_resiliency") is not None else "Unknown")
                ),
            ],
            [
                "Metrics Disabled",
                (
                    "Yes"
                    if cluster_info.get("disable_metrics")
                    else ("No" if cluster_info.get("disable_metrics") is not None else "Unknown")
                ),
            ],
            ["Capacity-Base 10", capacity_format],
        ]

        table_elements = self.brand_compliance.create_vast_table(
            cluster_data, f"Cluster Name: {cluster_name}", ["Function", "Status"]
        )
        content.extend(table_elements)

        return content

    def _create_hardware_inventory(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
        rack_layout_key: Optional[str] = None,
    ) -> List[Any]:
        """Create VAST brand-compliant hardware inventory section."""
        content = []

        # Add section heading with VAST styling
        heading_elements = self.brand_compliance.create_vast_section_heading("Hardware Summary", level=1)
        content.extend(heading_elements)

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        styles = getSampleStyleSheet()
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Hardware Summary section provides comprehensive inventory and operational status of all physical hardware components within the VAST Data cluster. This section includes detailed information about storage capacity utilization, compute nodes (CNodes), data nodes (DNodes), and their respective hardware specifications, operational status, and physical rack positioning. The capacity metrics show both logical and physical storage utilization, enabling capacity planning and performance optimization. Hardware inventory data is essential for understanding cluster scale, identifying hardware failures, planning maintenance windows, and ensuring proper rack organization for optimal cooling and cable management.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        hardware = data.get("hardware_inventory", {})
        eboxes_for_note = hardware.get("eboxes") or {}
        if eboxes_for_note:
            content.append(
                Paragraph(
                    "For clusters with EBox enclosures, DBox U height is documented in the inventory table below; the Physical Rack Layout diagram uses EBox U height and does not show DBox positions.",
                    overview_style,
                )
            )
            content.append(Spacer(1, 8))

        # Extract hardware collections early (used by inventory table and switch placement)
        cboxes = hardware.get("cboxes") or {}
        cnodes = hardware.get("cnodes") or []
        dboxes = hardware.get("dboxes") or {}
        dnodes = hardware.get("dnodes") or []
        eboxes = hardware.get("eboxes") or {}
        switches = hardware.get("switches") or []

        # Add storage capacity section
        cluster_info = data.get("cluster_summary", {})
        if any(
            cluster_info.get(field) is not None
            for field in [
                "usable_capacity_tb",
                "free_usable_capacity_tb",
                "drr_text",
                "physical_space_tb",
                "physical_space_in_use_tb",
                "free_physical_space_tb",
                "physical_space_in_use_percent",
                "logical_space_tb",
                "logical_space_in_use_tb",
                "free_logical_space_tb",
                "logical_space_in_use_percent",
            ]
        ):
            # Storage capacity table
            storage_data = []

            # Determine capacity unit based on capacity_base_10 setting
            capacity_unit = "TB" if cluster_info.get("capacity_base_10", True) else "TiB"

            # Usable capacity section
            if cluster_info.get("usable_capacity_tb") is not None:
                storage_data.append(
                    [
                        "Usable Capacity",
                        f"{round(cluster_info.get('usable_capacity_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("free_usable_capacity_tb") is not None:
                storage_data.append(
                    [
                        "Free Usable Capacity",
                        f"{round(cluster_info.get('free_usable_capacity_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("drr_text") and cluster_info.get("drr_text") != "Unknown":
                storage_data.append(
                    [
                        "Data Reduction Ratio (DRR)",
                        cluster_info.get("drr_text", "Unknown"),
                    ]
                )

            # Physical space section
            if cluster_info.get("physical_space_tb") is not None:
                storage_data.append(
                    [
                        "Physical Space",
                        f"{round(cluster_info.get('physical_space_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("physical_space_in_use_tb") is not None:
                storage_data.append(
                    [
                        "Physical Space In Use",
                        f"{round(cluster_info.get('physical_space_in_use_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("free_physical_space_tb") is not None:
                storage_data.append(
                    [
                        "Free Physical Space",
                        f"{round(cluster_info.get('free_physical_space_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("physical_space_in_use_percent") is not None:
                storage_data.append(
                    [
                        "Physical Space In Use %",
                        f"{round(cluster_info.get('physical_space_in_use_percent', 0))}%",
                    ]
                )

            # Logical space section
            if cluster_info.get("logical_space_tb") is not None:
                storage_data.append(
                    [
                        "Logical Space",
                        f"{round(cluster_info.get('logical_space_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("logical_space_in_use_tb") is not None:
                storage_data.append(
                    [
                        "Logical Space In Use",
                        f"{round(cluster_info.get('logical_space_in_use_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("free_logical_space_tb") is not None:
                storage_data.append(
                    [
                        "Free Logical Space",
                        f"{round(cluster_info.get('free_logical_space_tb', 0))} {capacity_unit}",
                    ]
                )
            if cluster_info.get("logical_space_in_use_percent") is not None:
                storage_data.append(
                    [
                        "Logical Space In Use %",
                        f"{round(cluster_info.get('logical_space_in_use_percent', 0))}%",
                    ]
                )

            if storage_data:
                storage_table_elements = self.brand_compliance.create_vast_table(
                    storage_data, "Storage Capacity", ["Metric", "Value"]
                )
                content.extend(storage_table_elements)
                content.append(Spacer(1, 12))

            # Pre-calculate switch positions for rack diagram.
            self.switch_rack_name = None
            self.manual_rack_placements: Dict[str, list] = {}

            manual_placements = data.get("manual_switch_placements")
            if manual_placements:
                sw_by_name = {sw.get("name", sw.get("hostname", "")): sw for sw in switches} if switches else {}
                for idx, mp in enumerate(manual_placements, start=1):
                    rn = mp.get("rack_name", "Unknown")
                    u_pos = int(mp.get("u_position", 0))
                    self.switch_positions[idx] = u_pos
                    sw_data = dict(sw_by_name.get(mp.get("switch_name"), {}))
                    sw_data["u_position"] = u_pos
                    sw_data.setdefault("name", mp.get("switch_name", ""))
                    if mp.get("model_key"):
                        sw_data["model"] = mp["model_key"]
                    elif mp.get("model") and "model" not in sw_data:
                        sw_data["model"] = mp["model"]
                    if mp.get("height_u"):
                        sw_data["height_u"] = mp["height_u"]
                    self.manual_rack_placements.setdefault(rn, []).append(sw_data)
                self.switch_rack_name = sorted(self.manual_rack_placements.keys())[0]
                self.logger.info("Manual switch placement: %s", self.switch_positions)
            if (not manual_placements) and switches and len(switches) == 2:
                # --- Auto placement: cascade through racks (same logic for ebox: above ebox then below ebox) ---
                hw_cnodes = hardware.get("cnodes") or []
                hw_dboxes_raw = hardware.get("dboxes") or {}
                hw_eboxes_raw = hardware.get("eboxes") or {}
                racks_info_auto = data.get("racks", [])
                vms_rack_names_auto = {str(r.get("name")).strip() for r in racks_info_auto if r.get("name")}

                per_rack_cboxes: Dict[str, list] = {}
                for _cbox_name, cbox_data in cboxes.items():
                    rack_name = cbox_data.get("rack_name") or "Unknown"
                    if rack_name == "Unknown" and vms_rack_names_auto and len(vms_rack_names_auto) == 1:
                        rack_name = next(iter(vms_rack_names_auto))
                    rack_unit = cbox_data.get("rack_unit") or ""
                    if not rack_unit and hw_cnodes:
                        for cnode in hw_cnodes:
                            if cnode.get("cbox_id") == cbox_data.get("id"):
                                rp = cnode.get("rack_position")
                                if rp is not None:
                                    rack_unit = f"U{rp}"
                                    break
                    if not rack_unit:
                        continue
                    model = cbox_data.get("model") or "Unknown"
                    if model == "Unknown" and hw_cnodes:
                        for cnode in hw_cnodes:
                            if cnode.get("cbox_id") == cbox_data.get("id"):
                                model = cnode.get("model", cnode.get("box_vendor", "")) or "Unknown"
                                break
                    cbox_entry = {
                        "id": cbox_data.get("id"),
                        "model": model,
                        "rack_unit": rack_unit,
                        "state": cbox_data.get("state", "ACTIVE"),
                    }
                    per_rack_cboxes.setdefault(rack_name, []).append(cbox_entry)

                per_rack_dboxes: Dict[str, list] = {}
                for dbox_name, dbox_info in hw_dboxes_raw.items():
                    rack_unit = dbox_info.get("rack_unit", "")
                    rack_name = dbox_info.get("rack_name") or "Unknown"
                    if rack_unit:
                        dbox_entry = {
                            "id": dbox_info.get("id"),
                            "model": dbox_info.get("hardware_type", "Unknown"),
                            "rack_unit": rack_unit,
                            "state": dbox_info.get("state", "ACTIVE"),
                        }
                        per_rack_dboxes.setdefault(rack_name, []).append(dbox_entry)

                per_rack_eboxes: Dict[str, list] = {}
                for ebox_name, ebox_info in hw_eboxes_raw.items():
                    rack_unit = ebox_info.get("rack_unit", "")
                    rack_name = ebox_info.get("rack_name") or "Unknown"
                    if rack_unit:
                        ebox_entry = {
                            "id": ebox_info.get("id"),
                            "model": "ebox",
                            "hardware_type": "ebox",
                            "rack_unit": rack_unit,
                            "state": ebox_info.get("state", "ACTIVE"),
                        }
                        per_rack_eboxes.setdefault(rack_name, []).append(ebox_entry)

                all_rack_names = sorted(
                    set(list(per_rack_cboxes.keys()) + list(per_rack_dboxes.keys()) + list(per_rack_eboxes.keys()))
                )
                if vms_rack_names_auto:
                    all_rack_names = [r for r in all_rack_names if r in vms_rack_names_auto]

                for try_rack in all_rack_names:
                    rack_cb = per_rack_cboxes.get(try_rack, [])
                    rack_db = per_rack_dboxes.get(try_rack, [])
                    rack_ebox = per_rack_eboxes.get(try_rack, [])
                    if not rack_cb and not rack_db and not rack_ebox:
                        continue
                    temp_rack_gen = RackDiagram(
                        library_path=self.library_path,
                        user_images_dir=self.user_images_dir,
                    )
                    calculated_positions = temp_rack_gen._calculate_switch_positions(
                        rack_cb, rack_db, len(switches), switches=switches, eboxes=rack_ebox
                    )
                    if calculated_positions:
                        self.switch_positions = {idx: u_pos for idx, u_pos in enumerate(calculated_positions, start=1)}
                        self.switch_rack_name = try_rack
                        self.logger.info(
                            f"Switches assigned to rack '{try_rack}': " f"positions {self.switch_positions}"
                        )
                        break

                if self.switch_rack_name is None:
                    self.logger.warning(
                        "Auto switch placement failed for all racks — " "switches will not appear in rack diagrams"
                    )

        # Consolidated Hardware Inventory table with VAST styling
        if cboxes or dboxes or eboxes or switches:
            inventory_elements = self._create_consolidated_inventory_table(
                cboxes, cnodes, dboxes, dnodes, switches, eboxes
            )
            content.extend(inventory_elements)

        # Add Physical Rack Layout
        rack_positions = hardware.get("rack_positions_available", False)
        if rack_positions:
            content.append(PageBreak())

            # Note: Section heading will be combined with first rack heading
            # No separate heading needed here

            # Generate rack diagrams - one per rack
            try:
                # Only include racks that exist in VMS (avoids "Unknown" or phantom racks)
                racks_info = data.get("racks", [])
                vms_rack_names = {str(r.get("name")).strip() for r in racks_info if r.get("name")}

                # Group hardware by rack_name
                racks_data: dict[str, Any] = {}  # rack_name -> {cboxes: [], dboxes: [], eboxes: [], switches: []}

                # Get CBox information and group by rack (one entry per CBox, not per CNode)
                hw_cnodes = hardware.get("cnodes") or []
                for cbox_name, cbox_data in cboxes.items():
                    rack_name = cbox_data.get("rack_name") or "Unknown"
                    if rack_name == "Unknown" and vms_rack_names and len(vms_rack_names) == 1:
                        rack_name = next(iter(vms_rack_names))
                    rack_unit = cbox_data.get("rack_unit") or ""
                    if not rack_unit and hw_cnodes:
                        for cnode in hw_cnodes:
                            if cnode.get("cbox_id") == cbox_data.get("id"):
                                rp = cnode.get("rack_position")
                                if rp is not None:
                                    rack_unit = f"U{rp}"
                                    break
                    if not rack_unit:
                        continue
                    if rack_name not in racks_data:
                        racks_data[rack_name] = {
                            "cboxes": [],
                            "dboxes": [],
                            "eboxes": [],
                            "switches": [],
                        }
                    model = cbox_data.get("model") or "Unknown"
                    if model == "Unknown" and hw_cnodes:
                        for cnode in hw_cnodes:
                            if cnode.get("cbox_id") == cbox_data.get("id"):
                                model = cnode.get("model", cnode.get("box_vendor", "")) or "Unknown"
                                break
                    cbox_entry = {
                        "id": cbox_data.get("id"),
                        "name": cbox_name,
                        "model": model,
                        "rack_unit": rack_unit,
                        "state": cbox_data.get("state", "ACTIVE"),
                    }
                    racks_data[rack_name]["cboxes"].append(cbox_entry)

                # Get DBox information and group by rack
                hw_dboxes = hardware.get("dboxes", {})
                for dbox_name, dbox_info in hw_dboxes.items():
                    rack_unit = dbox_info.get("rack_unit", "")
                    rack_name = dbox_info.get("rack_name") or "Unknown"

                    if rack_unit:
                        if rack_name not in racks_data:
                            racks_data[rack_name] = {
                                "cboxes": [],
                                "dboxes": [],
                                "eboxes": [],
                                "switches": [],
                            }

                        dbox_data = {
                            "id": dbox_info.get("id"),
                            "name": dbox_name,
                            "model": dbox_info.get("hardware_type", "Unknown"),
                            "rack_unit": rack_unit,
                            "state": dbox_info.get("state", "ACTIVE"),
                        }
                        racks_data[rack_name]["dboxes"].append(dbox_data)

                # Get EBox information and group by rack (for ebox clusters)
                hw_eboxes = hardware.get("eboxes") or {}
                is_ebox_cluster = bool(hw_eboxes)
                for ebox_name, ebox_info in hw_eboxes.items():
                    rack_unit = ebox_info.get("rack_unit", "")
                    rack_name = ebox_info.get("rack_name") or "Unknown"
                    if rack_unit:
                        if rack_name not in racks_data:
                            racks_data[rack_name] = {
                                "cboxes": [],
                                "dboxes": [],
                                "eboxes": [],
                                "switches": [],
                            }
                        # EBox model from associated CNode box_vendor (for library image lookup)
                        ebox_id = ebox_info.get("id")
                        model_key = "ebox"
                        for cnode in hw_cnodes:
                            if cnode.get("ebox_id") == ebox_id:
                                bv = cnode.get("box_vendor") or ""
                                if isinstance(bv, str) and bv:
                                    model_key = bv.split(",")[0].strip()
                                break
                        ebox_data = {
                            "id": ebox_info.get("id"),
                            "name": ebox_info.get("name", ""),
                            "model": model_key,
                            "hardware_type": model_key,
                            "rack_unit": rack_unit,
                            "state": ebox_info.get("state", "ACTIVE"),
                        }
                        racks_data[rack_name]["eboxes"].append(ebox_data)

                # Assign switches to racks for diagram generation.
                switches = hardware.get("switches") or []
                manual_rack_map = getattr(self, "manual_rack_placements", {})

                if manual_rack_map:
                    # Manual placement: distribute switches to their specified racks
                    # with explicit rack_unit so generate_rack_diagram uses them directly.
                    # Build a case-insensitive lookup for rack name resolution.
                    racks_data_lower = {k.lower().strip(): k for k in racks_data}
                    assigned_any = False
                    for rn, placements in manual_rack_map.items():
                        target_rn = rn
                        if rn not in racks_data:
                            # Try case-insensitive match
                            resolved = racks_data_lower.get(rn.lower().strip())
                            if resolved:
                                self.logger.info(
                                    f"Manual switch placement rack '{rn}' resolved to '{resolved}' (case-insensitive)"
                                )
                                target_rn = resolved
                            elif len(racks_data) == 1:
                                # Single hardware rack — assign there with a warning
                                target_rn = next(iter(racks_data))
                                self.logger.warning(
                                    f"Manual switch placement rack '{rn}' not found in racks_data — "
                                    f"falling back to only available rack '{target_rn}'"
                                )
                            else:
                                self.logger.warning(
                                    f"Manual switch placement rack '{rn}' not found in racks_data "
                                    f"(available: {list(racks_data.keys())}). Skipping these switches."
                                )
                                continue
                        sw_list = []
                        for p in placements:
                            sw_list.append(
                                {
                                    "id": p.get("name", "Unknown"),
                                    "model": p.get("model", "switch"),
                                    "state": p.get("state", "ACTIVE"),
                                    "rack_unit": f"U{p['u_position']}",
                                }
                            )
                        racks_data[target_rn]["switches"] = sw_list
                        assigned_any = True
                        self.logger.info(f"Manual: assigned {len(sw_list)} switches to rack '{target_rn}'")
                    if not assigned_any:
                        self.logger.warning(
                            "Manual switch placement produced no assignments — falling back to auto-placement"
                        )
                _manual_ok = bool(manual_rack_map) and assigned_any if manual_rack_map else False
                if not _manual_ok and switches and hasattr(self, "switch_rack_name") and self.switch_rack_name:
                    # Auto placement: all switches go to the single winning rack.
                    switches_data = []
                    for switch in switches:
                        switches_data.append(
                            {
                                "id": switch.get("name", "Unknown"),
                                "model": switch.get("model", "switch"),
                                "state": switch.get("state", "ACTIVE"),
                            }
                        )
                    target_rack = self.switch_rack_name
                    if target_rack in racks_data:
                        racks_data[target_rack]["switches"] = switches_data
                        self.logger.info(f"Auto: assigned {len(switches_data)} switches to rack '{target_rack}'")
                    else:
                        self.logger.warning(f"Switch target rack '{target_rack}' not found in racks_data")

                # Build node_status_map for rack diagram status indicators
                hw_cnodes_all = hardware.get("cnodes") or []
                hw_dnodes_all = hardware.get("dnodes") or []
                hw_switches_inv = hardware.get("switches") or []

                cnodes_by_cbox: dict[int, list[dict[str, Any]]] = {}
                cnode_by_ebox: dict[int, dict[str, Any]] = {}
                dnodes_by_dbox: dict[int, list[dict[str, Any]]] = {}
                dnodes_by_ebox: dict[int, list[dict[str, Any]]] = {}
                switch_in_inventory: dict[str, str] = {}

                for cn in hw_cnodes_all:
                    cn_status = str(cn.get("status") or cn.get("state") or "ACTIVE").upper()
                    cn_is_mgmt = bool(cn.get("is_mgmt", False))
                    entry = {"status": cn_status, "is_mgmt": cn_is_mgmt}
                    cbox_id = cn.get("cbox_id")
                    ebox_id = cn.get("ebox_id")
                    if cbox_id is not None:
                        cnodes_by_cbox.setdefault(cbox_id, []).append(entry)
                    if ebox_id is not None:
                        cnode_by_ebox[ebox_id] = entry

                for dn in hw_dnodes_all:
                    dn_status = str(dn.get("status") or dn.get("state") or "ACTIVE").upper()
                    entry = {"status": dn_status}
                    dbox_id = dn.get("dbox_id")
                    ebox_id = dn.get("ebox_id")
                    if dbox_id is not None:
                        dnodes_by_dbox.setdefault(dbox_id, []).append(entry)
                    if ebox_id is not None:
                        dnodes_by_ebox.setdefault(ebox_id, []).append(entry)

                for sw in hw_switches_inv:
                    sw_name = str(sw.get("name") or "").strip()
                    sw_state = str(sw.get("state") or sw.get("status") or "ACTIVE").upper()
                    if sw_name:
                        switch_in_inventory[sw_name] = sw_state

                node_status_map: dict[str, Any] = {
                    "cnodes_by_cbox": cnodes_by_cbox,
                    "cnode_by_ebox": cnode_by_ebox,
                    "dnodes_by_dbox": dnodes_by_dbox,
                    "dnodes_by_ebox": dnodes_by_ebox,
                    "switch_in_inventory": switch_in_inventory,
                }

                # Generate one diagram per rack
                if racks_data:
                    # Build mapping of rack_name to rack_height_u from racks data if available
                    rack_height_map = {}  # rack_name -> rack_height_u
                    racks_info = data.get("racks", [])
                    if racks_info:
                        # Create mapping from rack_id to number_of_units
                        rack_id_to_height = {}
                        for rack in racks_info:
                            rack_id = rack.get("id")
                            number_of_units = rack.get("number_of_units")
                            if rack_id is not None and number_of_units is not None:
                                rack_id_to_height[rack_id] = number_of_units

                        # Map rack_name to rack_height_u by finding matching rack_id
                        # We need to match rack_name to rack_id - try to find it from cboxes/dboxes
                        for rack_name in racks_data.keys():
                            # Try to find rack_id from cboxes or dboxes
                            rack_id = None
                            for cbox_name, cbox_data in cboxes.items():
                                if cbox_data.get("rack_name") == rack_name:
                                    rack_id = cbox_data.get("rack_id")
                                    if rack_id is not None:
                                        break

                            if rack_id is None:
                                for dbox_name, dbox_data in hw_dboxes.items():
                                    if dbox_data.get("rack_name") == rack_name:
                                        rack_id = dbox_data.get("rack_id")
                                        if rack_id is not None:
                                            break

                            # Get rack height from mapping
                            if rack_id is not None and rack_id in rack_id_to_height:
                                rack_height_map[rack_name] = rack_id_to_height[rack_id]

                    # Sort racks by name; include only racks that exist in VMS (never include "Unknown")
                    sorted_racks = sorted(racks_data.keys())
                    if vms_rack_names:
                        sorted_racks = [r for r in sorted_racks if r in vms_rack_names]
                        self.logger.info(
                            "Physical Rack Layout: only including racks present in VMS (%s)",
                            sorted_racks,
                        )
                    else:
                        sorted_racks = [r for r in sorted_racks if r and str(r).strip() != "Unknown"]
                        if "Unknown" in racks_data:
                            self.logger.info("Physical Rack Layout: excluding 'Unknown' rack (not in VMS)")

                    for idx, rack_name in enumerate(sorted_racks):
                        rack_hw = racks_data[rack_name]
                        rack_cboxes = rack_hw["cboxes"]
                        rack_dboxes = rack_hw["dboxes"]
                        rack_eboxes = rack_hw.get("eboxes") or []
                        rack_switches = rack_hw["switches"] if rack_hw["switches"] else None

                        # Only generate diagram if rack has hardware (cboxes, dboxes, or eboxes)
                        if rack_cboxes or rack_dboxes or rack_eboxes:
                            # Get rack height for this rack (default to 42U if not found)
                            rack_height_u = rack_height_map.get(rack_name, 42)

                            # Create rack diagram generator with appropriate rack height
                            # Reserve headroom for the heading, spacer, and table padding
                            # that share the page with the diagram
                            fw = getattr(self, "_frame_width", 7.27 * inch)
                            fh = getattr(self, "_frame_height", 8.5 * inch)
                            rack_page_h = fh - 1.0 * inch
                            rack_gen = RackDiagram(
                                page_width=fw,
                                page_height=rack_page_h,
                                rack_height_u=rack_height_u,
                                library_path=self.library_path,
                                user_images_dir=self.user_images_dir,
                            )

                            # EBox clusters: use ebox U height in diagram; disregard dbox in diagram
                            diagram_dboxes = [] if is_ebox_cluster else rack_dboxes
                            diagram_eboxes = rack_eboxes if is_ebox_cluster else []

                            rack_drawing, switch_positions_map = rack_gen.generate_rack_diagram(
                                rack_cboxes,
                                diagram_dboxes,
                                rack_switches,
                                rack_name=rack_name,
                                eboxes=diagram_eboxes,
                                node_status_map=node_status_map,
                            )

                            # Add rack name heading before diagram
                            # First rack gets combined heading, subsequent racks get simple heading
                            rack_heading_style = ParagraphStyle(
                                "Rack_Heading",
                                parent=styles["Heading2"],
                                fontSize=14,
                                spaceAfter=8,
                                spaceBefore=(12 if idx > 0 else 0),  # No space before first heading
                                textColor=self.brand_compliance.colors.BACKGROUND_DARK,
                            )

                            if idx == 0:
                                # First rack: Combined heading with section title
                                heading_text = f"Physical Rack Layout - Rack: {rack_name}"
                                # Place page marker for Physical Rack Layout section immediately after first heading
                                if page_tracker is not None and rack_layout_key == "rack_layout":
                                    content.append(PageMarker("rack_layout", page_tracker))
                            else:
                                # Subsequent racks: Simple heading
                                heading_text = f"Rack: {rack_name}"

                            content.append(Paragraph(heading_text, rack_heading_style))
                            content.append(Spacer(1, 0.2 * inch))

                            # Center the rack diagram on the page using a table
                            from reportlab.platypus import Table as RLTable

                            fw = getattr(self, "_frame_width", 7.5 * inch)
                            rack_table = RLTable(
                                [[rack_drawing]],
                                colWidths=[fw],
                            )
                            rack_table.setStyle(
                                TableStyle(
                                    [
                                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                                    ]
                                )
                            )
                            content.append(rack_table)

                            # Add page break between racks (except for last one)
                            if rack_name != sorted_racks[-1]:
                                content.append(PageBreak())

                            switch_msg = f", {len(rack_switches)} Switches" if rack_switches else ""
                            self.logger.info(
                                f"Added rack diagram for {rack_name}: {len(rack_cboxes)} CBoxes, {len(rack_dboxes)} DBoxes{switch_msg}"
                            )
                else:
                    # Fallback to placeholder if no position data
                    layout_elements = self.brand_compliance.create_vast_2d_diagram_placeholder(
                        "Physical Rack Layout",
                        "Rack position data not available for this cluster.",
                    )
                    content.extend(layout_elements)
                    self.logger.warning("No rack position data available for diagram")

            except Exception as e:
                self.logger.error(f"Error generating rack diagram: {e}", exc_info=True)
                # Fallback to placeholder on error
                layout_elements = self.brand_compliance.create_vast_2d_diagram_placeholder(
                    "Physical Rack Layout",
                    "Visual representation of hardware positioning in the rack with U-number assignments.",
                )
                content.extend(layout_elements)

        return content

    def _create_network_configuration(self, data: Dict[str, Any]) -> List[Any]:
        """Create network configuration section."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        content.append(Paragraph("Network Configuration", heading_style))
        content.append(Spacer(1, 12))

        sections = data.get("sections", {})
        network_config = sections.get("network_configuration", {}).get("data", {})

        # DNS Configuration
        dns_config = network_config.get("dns")
        if dns_config:
            content.append(Paragraph("<b>DNS Configuration:</b>", normal_style))
            content.append(Paragraph(f"• Enabled: {dns_config.get('enabled', False)}", normal_style))
            servers = dns_config.get("servers", [])
            if servers:
                content.append(Paragraph(f"• Servers: {', '.join(servers)}", normal_style))
            search_domains = dns_config.get("search_domains", [])
            if search_domains:
                content.append(Paragraph(f"• Search Domains: {', '.join(search_domains)}", normal_style))
            content.append(Spacer(1, 8))

        # NTP Configuration
        ntp_config = network_config.get("ntp")
        if ntp_config:
            content.append(Paragraph("<b>NTP Configuration:</b>", normal_style))
            content.append(Paragraph(f"• Enabled: {ntp_config.get('enabled', False)}", normal_style))
            servers = ntp_config.get("servers", [])
            if servers:
                content.append(Paragraph(f"• Servers: {', '.join(servers)}", normal_style))
            content.append(Spacer(1, 8))

        # VIP Pools
        vippool_config = network_config.get("vippools")
        if vippool_config:
            pools = vippool_config.get("pools", [])
            pool_count = len(pools) if isinstance(pools, list) else 0
            content.append(Paragraph(f"<b>VIP Pools:</b> {pool_count} pools configured", normal_style))
            content.append(Spacer(1, 8))

        # Cluster Network Configuration
        cluster_summary = data.get("cluster_summary", {})
        if cluster_summary:
            content.append(Paragraph("<b>Cluster Network Configuration:</b>", normal_style))

            # Management and Gateway Configuration - Always show placeholders
            management_vips = cluster_summary.get("management_vips")
            management_vips_display = (
                management_vips if management_vips and management_vips != "Unknown" else "Not Configured"
            )
            content.append(Paragraph(f"• Management VIPs: {management_vips_display}", normal_style))

            external_gateways = cluster_summary.get("external_gateways")
            external_gateways_display = (
                external_gateways if external_gateways and external_gateways != "Unknown" else "Not Configured"
            )
            content.append(Paragraph(f"• External Gateways: {external_gateways_display}", normal_style))

            # DNS and NTP Configuration - Always show placeholders
            dns = cluster_summary.get("dns")
            dns_display = dns if dns and dns != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• DNS Server: {dns_display}", normal_style))

            ntp = cluster_summary.get("ntp")
            ntp_display = ntp if ntp and ntp != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• NTP Server: {ntp_display}", normal_style))

            # Network Interface Configuration - Always show placeholders
            ext_netmask = cluster_summary.get("ext_netmask")
            ext_netmask_display = ext_netmask if ext_netmask and ext_netmask != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• External Netmask: {ext_netmask_display}", normal_style))

            auto_ports_ext_iface = cluster_summary.get("auto_ports_ext_iface")
            auto_ports_ext_iface_display = (
                auto_ports_ext_iface if auto_ports_ext_iface and auto_ports_ext_iface != "Unknown" else "Not Configured"
            )
            content.append(
                Paragraph(
                    f"• Auto Ports External Interface: {auto_ports_ext_iface_display}",
                    normal_style,
                )
            )

            # IPMI Configuration - Always show placeholders
            b2b_ipmi = cluster_summary.get("b2b_ipmi")
            b2b_ipmi_display = b2b_ipmi if b2b_ipmi is not None else "Not Configured"
            content.append(Paragraph(f"• B2B IPMI: {b2b_ipmi_display}", normal_style))

            ipmi_gateway = cluster_summary.get("ipmi_gateway")
            ipmi_gateway_display = ipmi_gateway if ipmi_gateway and ipmi_gateway != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• IPMI Gateway: {ipmi_gateway_display}", normal_style))

            ipmi_netmask = cluster_summary.get("ipmi_netmask")
            ipmi_netmask_display = ipmi_netmask if ipmi_netmask and ipmi_netmask != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• IPMI Netmask: {ipmi_netmask_display}", normal_style))

            # MTU Configuration - Always show placeholders
            eth_mtu = cluster_summary.get("eth_mtu")
            eth_mtu_display = eth_mtu if eth_mtu is not None else "Not Configured"
            content.append(Paragraph(f"• Ethernet MTU: {eth_mtu_display}", normal_style))

            ib_mtu = cluster_summary.get("ib_mtu")
            ib_mtu_display = ib_mtu if ib_mtu is not None else "Not Configured"
            content.append(Paragraph(f"• InfiniBand MTU: {ib_mtu_display}", normal_style))

            nb_eth_mtu = cluster_summary.get("nb_eth_mtu")
            nb_eth_mtu_display = nb_eth_mtu if nb_eth_mtu is not None else "Not Configured"
            content.append(Paragraph(f"• NVMe/TCP Ethernet MTU: {nb_eth_mtu_display}", normal_style))

        return content

    def _create_cluster_network_configuration(self, data: Dict[str, Any]) -> List[Any]:
        """Create cluster-wide network configuration section."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        content.append(Paragraph("Cluster Network Configuration", heading_style))
        content.append(Spacer(1, 12))

        sections = data.get("sections", {})
        cluster_network_config = sections.get("cluster_network_configuration", {}).get("data", {})

        if cluster_network_config:
            # Management VIPs
            management_vips = cluster_network_config.get("management_vips", "Not Configured")
            if management_vips != "Not Configured":
                content.append(Paragraph(f"• Management VIPs: {management_vips}", normal_style))
            else:
                content.append(Paragraph("• Management VIPs: Not Configured", normal_style))

            # Management VIP (single)
            mgmt_vip = cluster_network_config.get("mgmt_vip", "Not Configured")
            if mgmt_vip != "Not Configured":
                content.append(Paragraph(f"• Management VIP: {mgmt_vip}", normal_style))

            # Management Inner VIP
            mgmt_inner_vip = cluster_network_config.get("mgmt_inner_vip", "Not Configured")
            if mgmt_inner_vip != "Not Configured":
                content.append(Paragraph(f"• Management Inner VIP: {mgmt_inner_vip}", normal_style))

            # Management Inner VIP CNode
            mgmt_inner_vip_cnode = cluster_network_config.get("mgmt_inner_vip_cnode", "Not Configured")
            if mgmt_inner_vip_cnode != "Not Configured":
                content.append(
                    Paragraph(
                        f"• Management Inner VIP CNode: {mgmt_inner_vip_cnode}",
                        normal_style,
                    )
                )

            # External Gateways
            external_gateways = cluster_network_config.get("external_gateways", "Not Configured")
            if external_gateways != "Not Configured":
                content.append(Paragraph(f"• External Gateways: {external_gateways}", normal_style))
            else:
                content.append(Paragraph("• External Gateways: Not Configured", normal_style))

            # DNS Servers
            dns = cluster_network_config.get("dns", "Not Configured")
            if dns != "Not Configured":
                content.append(Paragraph(f"• DNS Servers: {dns}", normal_style))
            else:
                content.append(Paragraph("• DNS Servers: Not Configured", normal_style))

            # NTP Servers
            ntp = cluster_network_config.get("ntp", "Not Configured")
            if ntp != "Not Configured":
                content.append(Paragraph(f"• NTP Servers: {ntp}", normal_style))
            else:
                content.append(Paragraph("• NTP Servers: Not Configured", normal_style))

            # Network Interface Configuration
            ext_netmask = cluster_network_config.get("ext_netmask", "Unknown")
            ext_netmask_display = ext_netmask if ext_netmask != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• External Netmask: {ext_netmask_display}", normal_style))

            auto_ports_ext_iface = cluster_network_config.get("auto_ports_ext_iface", "Unknown")
            auto_ports_ext_iface_display = (
                auto_ports_ext_iface if auto_ports_ext_iface != "Unknown" else "Not Configured"
            )
            content.append(
                Paragraph(
                    f"• Auto Ports External Interface: {auto_ports_ext_iface_display}",
                    normal_style,
                )
            )

            # IPMI Configuration
            b2b_ipmi = cluster_network_config.get("b2b_ipmi", False)
            content.append(Paragraph(f"• B2B IPMI: {b2b_ipmi}", normal_style))

            # MTU Configuration
            eth_mtu = cluster_network_config.get("eth_mtu", "Unknown")
            eth_mtu_display = eth_mtu if eth_mtu != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• Ethernet MTU: {eth_mtu_display}", normal_style))

            ib_mtu = cluster_network_config.get("ib_mtu", "Unknown")
            ib_mtu_display = ib_mtu if ib_mtu != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• InfiniBand MTU: {ib_mtu_display}", normal_style))

            nb_eth_mtu = cluster_network_config.get("nb_eth_mtu", "Unknown")
            nb_eth_mtu_display = nb_eth_mtu if nb_eth_mtu != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• NVMe/TCP Ethernet MTU: {nb_eth_mtu_display}", normal_style))

            # IPMI Gateway and Netmask
            ipmi_gateway = cluster_network_config.get("ipmi_gateway", "Unknown")
            ipmi_gateway_display = ipmi_gateway if ipmi_gateway != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• IPMI Gateway: {ipmi_gateway_display}", normal_style))

            ipmi_netmask = cluster_network_config.get("ipmi_netmask", "Unknown")
            ipmi_netmask_display = ipmi_netmask if ipmi_netmask != "Unknown" else "Not Configured"
            content.append(Paragraph(f"• IPMI Netmask: {ipmi_netmask_display}", normal_style))

        return content

    def _create_cnodes_network_configuration(self, data: Dict[str, Any]) -> List[Any]:
        """Create CNodes network configuration section with scale-out support."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        content.append(Paragraph("CNodes Network Configuration", heading_style))
        content.append(Spacer(1, 12))

        sections = data.get("sections", {})
        cnodes_network_config = sections.get("cnodes_network_configuration", {}).get("data", {})

        cnodes = cnodes_network_config.get("cnodes", [])
        total_cnodes = cnodes_network_config.get("total_cnodes", 0)

        if cnodes:
            content.append(Paragraph(f"<b>Total CNodes:</b> {total_cnodes}", normal_style))
            content.append(Spacer(1, 12))

            # Create table for CNodes with scale-out support
            table_data = [
                [
                    "ID",
                    "Hostname",
                    "Mgmt IP",
                    "IPMI IP",
                    "Box Vendor",
                    "VAST OS",
                    "VMS Host",
                    "TPM Support",
                    "Single NIC",
                    "Net Type",
                ]
            ]

            for cnode in cnodes:
                table_data.append(
                    [
                        str(cnode.get("id", "Unknown")),
                        cnode.get("hostname", "Unknown"),
                        cnode.get("mgmt_ip", "Unknown"),
                        cnode.get("ipmi_ip", "Unknown"),
                        (
                            cnode.get("box_vendor", "Unknown")[:30] + "..."
                            if len(cnode.get("box_vendor", "")) > 30
                            else cnode.get("box_vendor", "Unknown")
                        ),
                        cnode.get("vast_os", "Unknown"),
                        "Yes" if cnode.get("is_vms_host", False) else "No",
                        ("Yes" if cnode.get("tpm_boot_dev_encryption_supported", False) else "No"),
                        "Yes" if cnode.get("single_nic", False) else "No",
                        cnode.get("net_type", "Unknown"),
                    ]
                )

            page_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
            table = Table(
                table_data,
                colWidths=[
                    page_width * 0.08,  # ID
                    page_width * 0.18,  # Hostname
                    page_width * 0.12,  # Mgmt IP
                    page_width * 0.12,  # IPMI IP
                    page_width * 0.20,  # Box Vendor
                    page_width * 0.12,  # VAST OS
                    page_width * 0.06,  # VMS Host
                    page_width * 0.06,  # TPM Support
                    page_width * 0.06,  # Single NIC
                    page_width * 0.10,  # Net Type
                ],
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), self._font("bold")),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("FONTSIZE", (0, 1), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
                    ]
                )
            )

            content.append(table)
        else:
            content.append(Paragraph("No CNodes network configuration data available", normal_style))

        return content

    def _create_dnodes_network_configuration(self, data: Dict[str, Any]) -> List[Any]:
        """Create DNodes network configuration section with scale-out support."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        content.append(Paragraph("DNodes Network Configuration", heading_style))
        content.append(Spacer(1, 12))

        sections = data.get("sections", {})
        dnodes_network_config = sections.get("dnodes_network_configuration", {}).get("data", {})

        dnodes = dnodes_network_config.get("dnodes", [])
        total_dnodes = dnodes_network_config.get("total_dnodes", 0)

        if dnodes:
            content.append(Paragraph(f"<b>Total DNodes:</b> {total_dnodes}", normal_style))
            content.append(Spacer(1, 12))

            # Create table for DNodes with scale-out support
            table_data = [
                [
                    "ID",
                    "Hostname",
                    "Mgmt IP",
                    "IPMI IP",
                    "Box Vendor",
                    "VAST OS",
                    "Position",
                    "Ceres",
                    "Ceres v2",
                    "Net Type",
                ]
            ]

            for dnode in dnodes:
                table_data.append(
                    [
                        str(dnode.get("id", "Unknown")),
                        dnode.get("hostname", "Unknown"),
                        dnode.get("mgmt_ip", "Unknown"),
                        dnode.get("ipmi_ip", "Unknown"),
                        (
                            dnode.get("box_vendor", "Unknown")[:30] + "..."
                            if len(dnode.get("box_vendor", "")) > 30
                            else dnode.get("box_vendor", "Unknown")
                        ),
                        dnode.get("vast_os", "Unknown"),
                        dnode.get("position", "Unknown"),
                        "Yes" if dnode.get("is_ceres", False) else "No",
                        "Yes" if dnode.get("is_ceres_v2", False) else "No",
                        dnode.get("net_type", "Unknown"),
                    ]
                )

            page_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
            table = Table(
                table_data,
                colWidths=[
                    page_width * 0.08,  # ID
                    page_width * 0.18,  # Hostname
                    page_width * 0.12,  # Mgmt IP
                    page_width * 0.12,  # IPMI IP
                    page_width * 0.20,  # Box Vendor
                    page_width * 0.12,  # VAST OS
                    page_width * 0.06,  # Position
                    page_width * 0.06,  # Ceres
                    page_width * 0.06,  # Ceres v2
                    page_width * 0.10,  # Net Type
                ],
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, -1), self._font("bold")),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("FONTSIZE", (0, 1), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
                    ]
                )
            )

            content.append(table)
        else:
            content.append(Paragraph("No DNodes network configuration data available", normal_style))

        return content

    def _safe_table_value(self, value: Any, default: str = "Not Configured") -> str:
        """
        Safely convert any value to a string for use in table cells.
        Handles lists, None, and other types that could cause ReportLab errors.

        Args:
            value: Any value that might be placed in a table cell
            default: Default string to return if value is None/empty

        Returns:
            String safe for use in ReportLab table cells
        """
        if value is None:
            return default
        if isinstance(value, list):
            if not value:
                return default
            return ", ".join(str(v) for v in value)
        if isinstance(value, (dict,)):
            return str(value) if value else default
        value_str = str(value)
        if not value_str or value_str in ("Unknown", "null", "None"):
            return default
        return value_str

    def _ip_sort_key(self, node: Dict[str, Any]) -> Tuple[int, ...]:
        """
        Convert IP address to tuple for sorting (lowest to highest).

        Args:
            node: Node dictionary with mgmt_ip field

        Returns:
            Tuple of integers for IP sorting, or (999, 999, 999, 999) for invalid/unknown IPs
        """
        mgmt_ip = node.get("mgmt_ip", "Unknown")
        if mgmt_ip == "Unknown":
            return (999, 999, 999, 999)  # Sort Unknown to end
        try:
            # Split IP and convert each octet to int
            parts = mgmt_ip.split(".")
            if len(parts) == 4:
                return tuple(int(part) for part in parts)
            else:
                return (999, 999, 999, 999)  # Invalid IPs to end
        except (ValueError, AttributeError):
            return (999, 999, 999, 999)  # Invalid IPs to end

    def _create_comprehensive_network_configuration(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """Create comprehensive network configuration section consolidating all network data."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        subheading_style = ParagraphStyle(
            "Subsection_Heading",
            parent=styles["Heading2"],
            fontSize=self.config.heading_font_size - 2,
            spaceAfter=8,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        # Note: PageBreak is handled by main build function, no need to add here

        # Add section heading with VAST styling
        heading_elements = self.brand_compliance.create_vast_section_heading("Network Configuration", level=1)
        content.extend(heading_elements)

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Network Configuration section provides comprehensive documentation of all network-related settings and connectivity parameters for the VAST Data cluster. This section includes cluster-wide network configuration, individual node network settings for both compute nodes (CNodes) and data nodes (DNodes), and network service configurations such as DNS and NTP. The network configuration data is essential for understanding cluster connectivity, troubleshooting network issues, validating network security settings, and ensuring proper network segmentation. This information supports network administrators in maintaining optimal network performance and security posture for the storage infrastructure.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        sections = data.get("sections", {})

        # Add Network Configuration summary table (similar to Storage Capacity)
        cluster_network_config = sections.get("cluster_network_configuration", {}).get("data", {})
        if cluster_network_config:
            network_summary_data = []

            # Management and Gateway settings - use _safe_table_value for potential list values
            mgmt_vips = self._safe_table_value(cluster_network_config.get("management_vips"))
            if mgmt_vips != "Not Configured":
                network_summary_data.append(["Management VIPs", mgmt_vips])

            ext_gateways = self._safe_table_value(cluster_network_config.get("external_gateways"))
            if ext_gateways != "Not Configured":
                network_summary_data.append(["External Gateways", ext_gateways])

            # DNS and NTP settings - use _safe_table_value for potential list values
            dns_servers = self._safe_table_value(cluster_network_config.get("dns"))
            if dns_servers != "Not Configured":
                network_summary_data.append(["DNS Servers", dns_servers])

            ntp_servers = self._safe_table_value(cluster_network_config.get("ntp"))
            if ntp_servers != "Not Configured":
                network_summary_data.append(["NTP Servers", ntp_servers])

            # Network interface settings - use _safe_table_value for potential list values
            ext_netmask = self._safe_table_value(cluster_network_config.get("ext_netmask"))
            if ext_netmask != "Not Configured":
                network_summary_data.append(["External Netmask", ext_netmask])

            auto_ports = self._safe_table_value(cluster_network_config.get("auto_ports_ext_iface"))
            if auto_ports != "Not Configured":
                network_summary_data.append(["Auto Ports Ext Interface", auto_ports])

            # MTU settings
            if cluster_network_config.get("eth_mtu") and cluster_network_config.get("eth_mtu") != "Not Configured":
                network_summary_data.append(
                    [
                        "Ethernet MTU",
                        str(cluster_network_config.get("eth_mtu", "Not Configured")),
                    ]
                )
            if cluster_network_config.get("ib_mtu") and cluster_network_config.get("ib_mtu") != "Not Configured":
                network_summary_data.append(
                    [
                        "InfiniBand MTU",
                        str(cluster_network_config.get("ib_mtu", "Not Configured")),
                    ]
                )
            if (
                cluster_network_config.get("nb_eth_mtu")
                and cluster_network_config.get("nb_eth_mtu") != "Not Configured"
            ):
                network_summary_data.append(
                    [
                        "NVMe/TCP Ethernet MTU",
                        str(cluster_network_config.get("nb_eth_mtu", "Not Configured")),
                    ]
                )

            # IPMI settings (always show, even if Not Configured) - use _safe_table_value
            ipmi_gateway = self._safe_table_value(cluster_network_config.get("ipmi_gateway"))
            network_summary_data.append(["IPMI Gateway", ipmi_gateway])
            ipmi_netmask = self._safe_table_value(cluster_network_config.get("ipmi_netmask"))
            network_summary_data.append(["IPMI Netmask", ipmi_netmask])

            # B2B IPMI setting
            if cluster_network_config.get("b2b_ipmi") is not None:
                network_summary_data.append(["B2B IPMI", str(cluster_network_config.get("b2b_ipmi", False))])

            # Net Type setting - use _safe_table_value for potential list values
            net_type = self._safe_table_value(cluster_network_config.get("net_type"))
            if net_type != "Not Configured":
                network_summary_data.append(["Net Type", net_type])

            if network_summary_data:
                network_table_elements = self.brand_compliance.create_vast_table(
                    network_summary_data, "Network Configuration", ["Setting", "Value"]
                )
                content.extend(network_table_elements)
                content.append(Spacer(1, 16))

        # 1. CNodes Network Configuration
        # For EBox clusters, use hardware_inventory directly (from /api/v7/cnodes/)
        # For non-EBox clusters, try network_settings first then fallback
        hardware_inventory = data.get("hardware_inventory", {})
        eboxes = hardware_inventory.get("eboxes", {})
        is_ebox_cluster = bool(eboxes)

        cnodes = []
        if is_ebox_cluster:
            # EBox clusters: use hardware_inventory cnodes (from /api/v7/cnodes/)
            hw_cnodes = hardware_inventory.get("cnodes", [])
            for c in hw_cnodes:
                cnodes.append(
                    {
                        "id": c.get("id", "Unknown"),
                        "name": c.get("name", "Unknown"),
                        "mgmt_ip": c.get("mgmt_ip", "Unknown"),
                        "ipmi_ip": c.get("ipmi_ip", "Unknown"),
                        "vast_os": c.get("os_version", "Unknown"),
                        "is_vms_host": c.get("is_mgmt", False),
                    }
                )
        else:
            # Non-EBox clusters: try network_settings, fallback to hardware_inventory
            cnodes_network_config = sections.get("cnodes_network_configuration", {}).get("data", {})
            cnodes = cnodes_network_config.get("cnodes", [])
            if not cnodes:
                hw_cnodes = hardware_inventory.get("cnodes", [])
                for c in hw_cnodes:
                    cnodes.append(
                        {
                            "id": c.get("id", "Unknown"),
                            "name": c.get("name", "Unknown"),
                            "mgmt_ip": c.get("mgmt_ip", "Unknown"),
                            "ipmi_ip": c.get("ipmi_ip", "Unknown"),
                            "vast_os": c.get("os_version", "Unknown"),
                            "is_vms_host": c.get("is_mgmt", False),
                        }
                    )

        if cnodes:
            # Sort CNodes by Mgmt IP (lowest to highest)
            cnodes = sorted(cnodes, key=self._ip_sort_key)

            # Create table for CNodes with scale-out support
            headers = [
                "ID",
                "Name",
                "Mgmt IP",
                "IPMI IP",
                "VAST OS",
                "VMS Host",
            ]

            table_data = []
            for cnode in cnodes:
                # Use 'name' field; fall back to 'hostname' for backward compatibility
                cnode_name = cnode.get("name") or cnode.get("hostname", "Unknown")
                table_data.append(
                    [
                        cnode.get("id", "Unknown"),
                        cnode_name,
                        cnode.get("mgmt_ip", "Unknown"),
                        cnode.get("ipmi_ip", "Unknown"),
                        cnode.get("vast_os", "Unknown"),
                        str(cnode.get("is_vms_host", False)),
                    ]
                )

            # Create table with pagination support
            table_elements = self.brand_compliance.create_vast_hardware_table_with_pagination(
                table_data, "CNode Network Configuration", headers
            )
            content.extend(table_elements)
            content.append(Spacer(1, 16))

        # 2. DNodes Network Configuration
        # For EBox clusters, use hardware_inventory directly (from /api/v7/dnodes/)
        dnodes = []
        if is_ebox_cluster:
            # EBox clusters: use hardware_inventory dnodes (from /api/v7/dnodes/)
            hw_dnodes = hardware_inventory.get("dnodes", [])
            for d in hw_dnodes:
                # Position: "primary" (when empty) or "virtual"
                raw_pos = d.get("position") or ""
                pos = "virtual" if raw_pos == "virtual" else "primary"
                dnodes.append(
                    {
                        "id": d.get("id", "Unknown"),
                        "name": d.get("name", "Unknown"),
                        "mgmt_ip": d.get("mgmt_ip", "Unknown"),
                        "ipmi_ip": d.get("ipmi_ip", "Unknown"),
                        "vast_os": d.get("os_version", "Unknown"),
                        "position": pos,
                    }
                )
        else:
            # Non-EBox clusters: try network_settings, fallback to hardware_inventory
            dnodes_network_config = sections.get("dnodes_network_configuration", {}).get("data", {})
            dnodes = dnodes_network_config.get("dnodes", [])
            if not dnodes:
                hw_dnodes = hardware_inventory.get("dnodes", [])
                for d in hw_dnodes:
                    # Position: "primary" (when empty) or "virtual"
                    raw_pos = d.get("position") or ""
                    pos = "virtual" if raw_pos == "virtual" else "primary"
                    dnodes.append(
                        {
                            "id": d.get("id", "Unknown"),
                            "name": d.get("name", "Unknown"),
                            "mgmt_ip": d.get("mgmt_ip", "Unknown"),
                            "ipmi_ip": d.get("ipmi_ip", "Unknown"),
                            "vast_os": d.get("os_version", "Unknown"),
                            "position": pos,
                        }
                    )

        if dnodes:
            # Sort DNodes by Mgmt IP (lowest to highest)
            dnodes = sorted(dnodes, key=self._ip_sort_key)

            # Create table for DNodes with scale-out support
            headers = [
                "ID",
                "Name",
                "Mgmt IP",
                "IPMI IP",
                "VAST OS",
                "Position",
            ]

            table_data = []
            for dnode in dnodes:
                # Use 'name' field; fall back to 'hostname' for backward compatibility
                dnode_name = dnode.get("name") or dnode.get("hostname", "Unknown")
                # Position: "primary" or "virtual"
                raw_pos = dnode.get("position") or ""
                position = "virtual" if raw_pos == "virtual" else "primary"
                table_data.append(
                    [
                        dnode.get("id", "Unknown"),
                        dnode_name,
                        dnode.get("mgmt_ip", "Unknown"),
                        dnode.get("ipmi_ip", "Unknown"),
                        dnode.get("vast_os", "Unknown"),
                        position,
                    ]
                )

            # Create table with pagination support
            table_elements = self.brand_compliance.create_vast_hardware_table_with_pagination(
                table_data, "DNode Network Configuration", headers
            )
            content.extend(table_elements)
            content.append(Spacer(1, 16))

        return content

    def _create_logical_network_diagram(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """
        Create logical network diagram section.

        Args:
            data: Processed cluster data
            page_tracker: Optional dictionary to capture page numbers
            section_key: Optional section key for page tracking

        Returns:
            List of reportlab flowables for the section
        """
        content = []

        # Add section heading with VAST styling
        heading_elements = self.brand_compliance.create_vast_section_heading("Logical Network Diagram", level=1)
        content.extend(heading_elements)

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        styles = getSampleStyleSheet()
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Logical Network Diagram provides a visual "
                "representation of the cluster's network topology, "
                "illustrating the connectivity between compute nodes "
                "(CBoxes), data nodes (DBoxes), network switches, "
                "and the customer network. This diagram shows the "
                "redundant network paths, switch interconnections, "
                "and how data flows through the storage infrastructure. "
                "Understanding the logical network topology "
                "is essential for network planning, troubleshooting "
                "connectivity issues, validating redundancy "
                "configurations, and ensuring optimal network performance "
                "across the storage cluster.",
                overview_style,
            )
        )
        content.append(Spacer(1, 16))

        # Determine diagram mode from config
        diagram_mode = self.config.network_diagram.get("mode", "detailed")

        # Prepare shared data for either renderer
        port_mapping_section = data.get("sections", {}).get("port_mapping", {})
        port_mapping_data = port_mapping_section.get("data", {}) if isinstance(port_mapping_section, dict) else {}
        hardware_inventory = data.get("hardware_inventory", {})

        cboxes_data = hardware_inventory.get("cboxes") or {}
        dboxes_data = hardware_inventory.get("dboxes") or {}
        eboxes_data = hardware_inventory.get("eboxes") or {}
        switches_data = hardware_inventory.get("switches") or []

        cboxes_list = list(cboxes_data.values()) if isinstance(cboxes_data, dict) else cboxes_data
        dboxes_list = list(dboxes_data.values()) if isinstance(dboxes_data, dict) else dboxes_data
        eboxes_list = list(eboxes_data.values()) if isinstance(eboxes_data, dict) else eboxes_data
        switches_list = switches_data if isinstance(switches_data, list) else []

        cnodes_data = hardware_inventory.get("cnodes") or {}
        dnodes_data = hardware_inventory.get("dnodes") or {}
        cnodes_list = list(cnodes_data.values()) if isinstance(cnodes_data, dict) else cnodes_data
        dnodes_list = list(dnodes_data.values()) if isinstance(dnodes_data, dict) else dnodes_data

        hardware_data = {
            "cboxes": cboxes_list,
            "dboxes": dboxes_list,
            "eboxes": eboxes_list,
            "switches": switches_list,
            "cnodes": cnodes_list,
            "dnodes": dnodes_list,
        }

        self.logger.info(
            f"Hardware data for diagram: {len(hardware_data['cboxes'])} CBoxes, "
            f"{len(hardware_data['dboxes'])} DBoxes, {len(hardware_data['eboxes'])} EBoxes, "
            f"{len(hardware_data['switches'])} Switches"
        )

        diagrams_dir = get_data_dir() / "output" / "diagrams"
        diagrams_dir.mkdir(parents=True, exist_ok=True)

        generated_paths: list = []

        # ---------- Detailed mode (rack-centric SVG) ----------
        if diagram_mode == "detailed":
            try:
                from network_diagram_v2 import create_rack_centric_diagram_generator

                gen = create_rack_centric_diagram_generator(
                    config=self.config.network_diagram,
                    assets_path=str(get_bundle_dir() / "assets"),
                    library_path=self.library_path,
                    user_images_dir=self.user_images_dir,
                )
                generated_paths = gen.generate(
                    port_mapping_data=port_mapping_data,
                    hardware_data=hardware_data,
                    output_dir=str(diagrams_dir),
                )
                self.logger.info("Detailed diagram generated %d page(s)", len(generated_paths))
            except Exception as e:
                self.logger.warning("Detailed diagram failed, falling back to compact: %s", e)
                diagram_mode = "compact"

        # ---------- Compact mode (legacy ReportLab) ----------
        if diagram_mode == "compact" or not generated_paths:
            try:
                from network_diagram import NetworkDiagramGenerator

                diagram_generator = NetworkDiagramGenerator(
                    assets_path=str(get_bundle_dir() / "assets"),
                    library_path=self.library_path,
                    user_images_dir=self.user_images_dir,
                )
                diagram_path = diagrams_dir / "network_topology.pdf"
                target_diagram_width = 6.5 * inch
                target_diagram_height = 6.0 * inch

                gen_path = diagram_generator.generate_network_diagram(
                    port_mapping_data=port_mapping_data,
                    hardware_data=hardware_data,
                    output_path=str(diagram_path),
                    drawing_size=(target_diagram_width, target_diagram_height),
                )
                if gen_path and Path(gen_path).exists():
                    generated_paths = [gen_path]
            except Exception as e:
                self.logger.error("Error generating compact network diagram: %s", e, exc_info=True)

        # Generate network diagram dynamically
        try:
            # Add color legend
            legend_style = ParagraphStyle(
                "Diagram_Legend",
                parent=styles["Normal"],
                fontSize=self.config.font_size - 1,
                textColor=colors.HexColor("#666666"),
                alignment=TA_CENTER,
                spaceAfter=12,
            )
            content.append(
                Paragraph(
                    "<font color='#0F9D58'><b>■</b> Green</font> = Network A | "
                    "<font color='#4285F4'><b>■</b> Blue</font> = Network B | "
                    "<font color='#7B1FA2'><b>■</b> Purple</font> = IPL/MLAG | "
                    "<font color='#757575'><b>■</b> Gray</font> = Spine",
                    legend_style,
                )
            )
            content.append(Spacer(1, 12))

            if generated_paths:
                from PIL import Image as PILImage

                for gp in generated_paths:
                    if not Path(gp).exists():
                        continue
                    self.logger.info(f"Network diagram generated: {gp}")

                    available_width = getattr(self, "_frame_width", 7.5 * inch)
                    available_height = getattr(self, "_frame_height", 10.19 * inch)

                    orientation = self.config.network_diagram.get("orientation", "portrait")
                    use_landscape = orientation == "landscape"
                    if use_landscape and diagram_mode == "detailed":
                        available_width = 10.0 * inch
                        available_height = 7.0 * inch

                    max_diagram_height = available_height - 1.5 * inch

                    with PILImage.open(str(gp)) as pil_img:
                        img_width, img_height = pil_img.size
                        aspect_ratio = img_width / img_height

                    target_width = available_width * 0.95
                    target_height = target_width / aspect_ratio
                    if target_height > max_diagram_height:
                        target_height = max_diagram_height
                        target_width = target_height * aspect_ratio

                    self.logger.info(
                        f"Network diagram sizing: original={img_width}x{img_height}, "
                        f"target={target_width:.1f}x{target_height:.1f}, "
                        f"available={available_width:.1f}x{max_diagram_height:.1f}"
                    )

                    img = Image(str(gp), width=target_width, height=target_height)

                    from reportlab.platypus import Table as RLTable
                    image_table = RLTable(
                        [[img]],
                        colWidths=[available_width],
                    )
                    image_table.setStyle(
                        TableStyle(
                            [
                                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ]
                        )
                    )

                    if use_landscape and diagram_mode == "detailed":
                        content.append(NextPageTemplate("landscape"))
                        content.append(PageBreak())

                    content.append(image_table)

                    if use_landscape and diagram_mode == "detailed":
                        content.append(NextPageTemplate("VastPage"))
                        content.append(PageBreak())

                self.logger.info("Embedded network diagram (%s mode)", diagram_mode)
                return content

            else:
                self.logger.warning("Network diagram PNG not available")

        except Exception as e:
            self.logger.error(f"Error generating network diagram: {e}", exc_info=True)

        # Show "No switch data available" message instead of placeholder image
        content.append(Spacer(1, 24))
        content.append(
            Paragraph(
                "<i>No switch data available. The network diagram could not be generated "
                "due to insufficient switch connectivity information.</i>",
                styles["Normal"],
            )
        )
        self.logger.info("Network diagram not generated - no switch data available")

        return content

    def _classify_port_purpose(self, port_name: str, speed: str, total_switches: int) -> str:
        """
        Classify port purpose based on speed, port number, and cluster topology.

        Args:
            port_name: Port name (e.g., "swp1", "eth1/1")
            speed: Port speed (e.g., "200G", "100G", None)
            total_switches: Total number of switches in cluster

        Returns:
            Port classification string
        """
        # Extract port number from name (handles swp1, eth1/1, etc.)
        try:
            port_num = int("".join(filter(str.isdigit, port_name)))
        except (ValueError, TypeError):
            port_num = 0

        # Classification logic based on speed and port position
        if speed == "200G":
            # 200G ports are typically data plane connections
            if 1 <= port_num <= 14:
                return "Data Plane (CNode)"
            elif 15 <= port_num <= 28:
                return "Data Plane (DNode)"
            else:
                return "Data Plane"

        elif speed == "100G":
            # 100G ports are typically IPLs, uplinks, or standard data plane
            if port_num in [29, 30]:
                # Ports 29-30 typically used for MLAG IPL (Inter-Peer Link)
                return "IPL (Inter-Peer Link)" if total_switches > 1 else "Reserved"
            elif port_num in [31, 32]:
                # Ports 31-32 typically used for uplinks to spine/core
                return "Uplink (Spine/Core)"
            else:
                return "Data Plane"

        elif speed == "Unconfigured" or not speed:
            # Unconfigured ports are unused or reserved
            return "Unused/Reserved"

        else:
            # Other speeds (40G, 10G, etc.)
            return "Data Plane"

    def _create_vnetmap_topology_tables(
        self,
        port_map: List[Dict[str, Any]],
        switches: List[Dict[str, Any]],
        styles: Any,
    ) -> List[Any]:
        """Render vnetmap 'Full Topology' and per-switch topology tables."""
        content: List[Any] = []

        unique_nodes = {e.get("node_hostname", "") for e in port_map if e.get("node_hostname")}
        unique_switches = {e.get("switch_ip", "") for e in port_map if e.get("switch_ip")}

        summary_style = ParagraphStyle(
            "VnetmapSummary",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=6,
            spaceBefore=4,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
        )
        content.append(
            Paragraph(
                f"<b>Topology Discovery:</b> Mapping nodes <b>{len(unique_nodes)}</b> "
                f"Switches <b>{len(unique_switches)}</b>",
                summary_style,
            )
        )
        content.append(Spacer(1, 8))

        # --- Full Topology table ---
        # Proportional weights: Node(wide) | Switch IP | Port(narrow) | Data IP | Interface | MAC | Net(narrow)
        full_headers = ["Node", "Switch IP", "Port", "Data IP", "Interface", "MAC", "Net"]
        full_col_weights = [5, 3, 1.2, 2.5, 2, 3.5, 0.8]
        full_data = []
        for e in sorted(port_map, key=lambda x: (x.get("node_hostname", ""), x.get("network", ""))):
            full_data.append([
                e.get("node_hostname", ""),
                e.get("switch_ip", ""),
                e.get("port", ""),
                e.get("node_ip", ""),
                e.get("interface", ""),
                e.get("mac", "") or "",
                e.get("network", ""),
            ])

        full_topo_elements = self.brand_compliance.create_vast_table(
            full_data, "Full Topology", full_headers,
            col_widths=full_col_weights, compact=True,
        )
        content.extend(full_topo_elements)

        # --- Per-switch topology tables ---
        # Proportional weights: Node(wide) | Port(narrow) | Data IP | Interface | MAC | Net(narrow)
        sw_col_weights = [5, 1.2, 2.5, 2, 3.5, 0.8]

        switch_groups: Dict[str, List[Dict[str, Any]]] = {}
        for e in port_map:
            sip = e.get("switch_ip", "")
            if sip:
                switch_groups.setdefault(sip, []).append(e)

        leaf_ips = sorted(switch_groups.keys())
        for switch_ip in leaf_ips:
            entries = switch_groups[switch_ip]
            networks = sorted({e.get("network", "?") for e in entries})
            subnets = sorted({e.get("node_ip", "").rsplit(".", 1)[0] for e in entries if e.get("node_ip")})
            net_label = ", ".join(networks) if networks else "?"

            sw_headers = ["Node", "Port", "Data IP", "Interface", "MAC", "Net"]
            sw_data = []
            for e in sorted(entries, key=lambda x: (x.get("node_hostname", ""), x.get("port", ""))):
                sw_data.append([
                    e.get("node_hostname", ""),
                    e.get("port", ""),
                    e.get("node_ip", ""),
                    e.get("interface", ""),
                    e.get("mac", "") or "",
                    e.get("network", ""),
                ])

            subnet_display = ", ".join(subnets) if subnets else ""
            sw_title = (
                f"Switch {switch_ip} — subnet {{{subnet_display}}}, network {{{net_label}}}"
            )
            sw_elements = self.brand_compliance.create_vast_table(
                sw_data, sw_title, sw_headers,
                col_widths=sw_col_weights, compact=True,
            )
            content.extend(sw_elements)

        content.append(Spacer(1, 8))
        return content

    def _create_port_mapping_section(
        self,
        port_mapping_data: Dict[str, Any],
        switches: List[Dict[str, Any]],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """
        Create port mapping section with detailed port-to-device mappings.

        Args:
            port_mapping_data: Port mapping data from data extractor
            switches: List of switch hardware data
            page_tracker: Optional dictionary to capture page numbers
            section_key: Optional section key for page tracking

        Returns:
            List of ReportLab elements for port mapping section
        """
        content = []
        styles = getSampleStyleSheet()

        # Add page break to ensure Port Mapping starts at top of new page
        content.append(PageBreak())

        # Add section heading
        heading_elements = self.brand_compliance.create_vast_section_heading("Port Mapping", level=1)
        content.extend(heading_elements)

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Port Mapping section provides detailed connectivity information showing which switch ports connect to which cluster nodes. This information is critical for troubleshooting network issues, validating cabling, and planning maintenance activities. The mapping uses standardized designations to clearly identify each connection point on both the switch side and node side.",
                overview_style,
            )
        )
        if port_mapping_data.get("partial") and port_mapping_data.get("partial_reason"):
            content.append(
                Paragraph(
                    f"<b>Note:</b> {port_mapping_data['partial_reason']}",
                    ParagraphStyle(
                        "PartialNote",
                        parent=overview_style,
                        backColor=self.brand_compliance.colors.VAST_BLUE_LIGHTEST,
                        borderPadding=6,
                    ),
                )
            )
        content.append(Spacer(1, 12))

        # Get port map organized by switch
        port_map = port_mapping_data.get("port_map", [])
        has_cross_connections = port_mapping_data.get("has_cross_connections", False)

        if not port_map:
            content.append(Paragraph("No port mapping data available", styles["Normal"]))
            return content

        # --- Vnetmap Topology Detail ---
        if port_mapping_data.get("data_source", "").startswith("vnetmap"):
            content.extend(self._create_vnetmap_topology_tables(port_map, switches, styles))

        # Group ports by switch
        ports_by_switch: dict[str, Any] = {}
        for entry in port_map:
            switch_ip = entry["switch_ip"]
            if switch_ip not in ports_by_switch:
                ports_by_switch[switch_ip] = []
            ports_by_switch[switch_ip].append(entry)

        # Sort ports within each switch by port number
        for switch_ip in ports_by_switch:
            ports_by_switch[switch_ip].sort(
                key=lambda x: (
                    int("".join(filter(str.isdigit, x["port"]))) if any(c.isdigit() for c in x["port"]) else 0
                )
            )

        # Only number switches that actually appear in the port map (leaf switches)
        port_map_switch_ips = {e["switch_ip"] for e in port_map if e.get("switch_ip")}
        leaf_switches = [
            sw for sw in switches if sw.get("mgmt_ip") in port_map_switch_ips
        ]
        if not leaf_switches:
            leaf_switches = list(switches)
        leaf_switches.sort(key=lambda s: s.get("mgmt_ip", ""))

        switch_ip_to_number = {}
        switch_port_speed_lookup: dict[str, dict[str, str]] = {}
        for idx, switch in enumerate(leaf_switches, start=1):
            mgmt_ip = switch.get("mgmt_ip")
            switch_ip_to_number[mgmt_ip] = idx
            speed_map: dict[str, str] = {}
            for port_info in switch.get("ports", []):
                port_name = port_info.get("name", "")
                port_speed = port_info.get("speed", "")
                if port_name and port_speed:
                    speed_map[port_name] = port_speed
                    if port_name.startswith("Eth") and "/" in port_name:
                        parts = port_name.split("/")
                        if len(parts) == 2:
                            speed_map[f"swp{parts[1]}"] = port_speed
                        elif len(parts) == 3:
                            speed_map[f"swp{parts[1]}/{parts[2]}"] = port_speed
                            parent_key = f"Eth{parts[0]}/{parts[1]}"
                            if parent_key not in speed_map:
                                speed_map[parent_key] = port_speed
            if mgmt_ip:
                switch_port_speed_lookup[mgmt_ip] = speed_map

        # Sort switches by switch number (Switch 1 before Switch 2)
        sorted_switch_ips = sorted(ports_by_switch.keys(), key=lambda ip: switch_ip_to_number.get(ip, 999))

        # Create port map table for each switch (in order: Switch 1, Switch 2, etc.)
        for switch_ip in sorted_switch_ips:
            connections = ports_by_switch[switch_ip]
            switch_num = switch_ip_to_number.get(switch_ip, "?")

            # Build table data - filter to show only primary physical connections
            table_data = []
            headers = ["Switch Port", "Node Connection", "Network", "Speed", "Notes"]

            # Track which ports we've already added to avoid duplicates
            # Key: (switch_designation, node_designation, network)
            seen_connections = set()

            for conn in connections:
                # Smart filtering: show one primary interface per network per node
                # CNodes: f0 = Net A (primary), f1 = Net B (primary), f2/f3 = bonded
                # DNodes: f0 = Net A (primary), f2/f3 = Net B (primary - different than CNodes!)

                interface = conn.get("interface", "")
                network = conn.get("network", "?")
                node_designation = conn.get("node_designation", "Unknown")

                # Determine if this is a DNode or CNode
                is_dnode = "DN" in node_designation
                is_cnode = "CN" in node_designation
                is_unknown = "UNKNOWN" in node_designation.upper()

                # Primary interface logic (simplified):
                # Show physical interfaces (f0, f1, f2, f3) - these are the primary physical ports
                # Skip bonded/virtual interfaces (bonds, VLANs, etc.)
                #
                # Network assignment (A or B) is already correctly determined
                # by which switch the connection is on, so we don't need to
                # make assumptions about which interface goes to which network.

                is_primary = False
                if "f0" in interface or "f1" in interface or "f2" in interface or "f3" in interface:
                    # This is a primary physical interface
                    is_primary = True

                # Skip non-primary interfaces (bonds, VLANs, etc.)
                if not is_primary:
                    continue

                # Use enhanced switch designation (e.g., SWA-P20)
                port_display = conn.get("switch_designation", conn["port"])

                # Use node designation (already have it from above)
                node_display = node_designation

                # Create unique key for this connection
                conn_key = (port_display, node_designation, network)

                # Skip if we've already added this connection
                if conn_key in seen_connections:
                    continue
                seen_connections.add(conn_key)

                port_name_raw = conn.get("port", "")
                speed = switch_port_speed_lookup.get(switch_ip, {}).get(
                    port_name_raw, switch_port_speed_lookup.get(switch_ip, {}).get(
                        conn.get("original_port", ""), ""
                    )
                ) or "Unknown"

                # Use notes from port map entry
                # Check if this is an EBox cluster entry (has ebox_id or ebox_node_type)
                is_ebox_entry = conn.get("ebox_id") is not None or conn.get("ebox_node_type") is not None

                if is_ebox_entry:
                    # EBox clusters: notes contain the CNode/DNode name
                    notes_str = conn.get("notes", "")
                    if not notes_str:
                        # Fallback for EBox entries without notes
                        notes_str = conn.get("node_name", "")
                else:
                    # Standard CBox/DBox clusters: show "Primary" for all active connections
                    notes_str = "Primary"

                table_data.append([port_display, node_display, network, speed, notes_str])

            # Add IPL/MLAG connections to this switch's table
            ipl_connections = port_mapping_data.get("ipl_connections", [])
            if ipl_connections:
                for ipl_conn in ipl_connections:
                    sw_des = ipl_conn.get("switch_designation", "")
                    peer_des = ipl_conn.get("node_designation", "")
                    ipl_note = ipl_conn.get("notes", "IPL")

                    if switch_num == 1:
                        table_data.append([
                            sw_des or "SWA",
                            peer_des or "SWB",
                            "A/B",
                            "",
                            ipl_note,
                        ])
                    elif switch_num == 2:
                        table_data.append([
                            peer_des or "SWB",
                            sw_des or "SWA",
                            "A/B",
                            "",
                            ipl_note,
                        ])

            # Create table
            table_title = f"Switch {switch_num} Port-to-Device Mapping"
            table_elements = self.brand_compliance.create_vast_table(table_data, table_title, headers)
            content.extend(table_elements)
            content.append(Spacer(1, 12))

        # Add diagnostic summary if available
        diagnostic_summary = port_mapping_data.get("diagnostic_summary")
        if diagnostic_summary:
            content.append(Spacer(1, 12))
            diagnostic_style = ParagraphStyle(
                "Diagnostic_Summary",
                parent=styles["Normal"],
                fontSize=self.config.font_size - 1,
                textColor=self.brand_compliance.colors.BACKGROUND_DARK,
                spaceAfter=8,
                leftIndent=12,
            )

            diagnostic_text = (
                f"<b>Collection Summary:</b><br/>"
                f"• Nodes with MACs collected: {diagnostic_summary.get('nodes_collected', 0)}<br/>"
                f"• CNode connections: {diagnostic_summary.get('cnode_connections', 0)}<br/>"
                f"• DNode connections: {diagnostic_summary.get('dnode_connections', 0)}<br/>"
                f"• Network A connections: {diagnostic_summary.get('network_a_connections', 0)}<br/>"
                f"• Network B connections: {diagnostic_summary.get('network_b_connections', 0)}<br/>"
                f"• Total MACs in switch tables: {diagnostic_summary.get('total_switch_macs', 0)}<br/>"
                f"• Switches queried: {diagnostic_summary.get('switches_queried', 0)}"
            )
            content.append(Paragraph(diagnostic_text, diagnostic_style))
            content.append(Spacer(1, 12))

        # Note: Cross-connection detection disabled - VAST dual-network design
        # naturally has both Network A and B on both switches for redundancy
        # Add cross-connection summary if any issues detected
        if False and has_cross_connections:  # Disabled
            content.append(Spacer(1, 12))

            warning_heading = self.brand_compliance.create_vast_section_heading(
                "⚠️ Network Configuration Issues Detected", level=2
            )
            content.extend(warning_heading)

            cross_connections = port_mapping_data.get("cross_connections", [])
            cross_summary = port_mapping_data.get("cross_connection_summary", "")

            issue_style = ParagraphStyle(
                "Issue_Style",
                parent=styles["Normal"],
                fontSize=self.config.font_size,
                textColor=colors.HexColor("#d62728"),  # Red warning color
                spaceAfter=8,
                spaceBefore=8,
                leftIndent=20,
                rightIndent=20,
            )

            content.append(
                Paragraph(
                    f"<b>Summary:</b> {cross_summary}",
                    issue_style,
                )
            )
            content.append(Spacer(1, 8))

            # List specific cross-connection issues
            issue_list_data = []
            issue_headers = [
                "Switch Port",
                "Node",
                "Actual Network",
                "Expected Network",
            ]

            for issue in cross_connections:
                issue_list_data.append(
                    [
                        issue.get("port", "Unknown"),
                        issue.get("node", issue.get("node_designation", "Unknown")),
                        f"Network {issue.get('actual_network', '?')}",
                        f"Network {issue.get('expected_network', '?')}",
                    ]
                )

            issue_table_elements = self.brand_compliance.create_vast_table(
                issue_list_data, "Cross-Connection Details", issue_headers
            )
            content.extend(issue_table_elements)
            content.append(Spacer(1, 12))

        # IPL/MLAG ports are now integrated into switch port tables above
        # No separate section needed

        return content

    def _create_switch_configuration(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """Create switch configuration section with port details."""
        content = []

        # Add page break to ensure Switch Configuration starts at top of new page
        content.append(PageBreak())

        # Add section heading with VAST styling
        heading_elements = self.brand_compliance.create_vast_section_heading("Switch Configuration", level=1)
        content.extend(heading_elements)

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        styles = getSampleStyleSheet()
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Switch Configuration section provides detailed information about the network switches that form the fabric interconnecting the VAST cluster nodes. This section documents switch hardware specifications, port configurations, operational status, and connectivity details. Understanding the switch topology is critical for network troubleshooting, capacity planning, and validating proper network segmentation. The port-level details enable network administrators to trace physical connectivity, identify unused ports, and plan for cluster expansion.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        # Get switch data
        hardware = data.get("hardware_inventory", {})
        switches = hardware.get("switches") or []

        if not switches:
            content.append(Paragraph("No switch data available", styles["Normal"]))
            return content

        # For each switch, create a detailed port configuration table on separate page
        for switch_num, switch in enumerate(switches, start=1):
            # Add page break before each switch (except the first)
            if switch_num > 1:
                content.append(PageBreak())

            # Get switch name first to use in consolidated heading
            switch_name = switch.get("name", "Unknown")

            hostname = switch.get("hostname", "Unknown")
            model = switch.get("model", "Unknown")
            serial = switch.get("serial", "Unknown")
            firmware_version = switch.get("firmware_version", "Unknown")
            mgmt_ip = switch.get("mgmt_ip", "Unknown")
            switch_type = switch.get("switch_type", "Unknown")
            state = switch.get("state", "Unknown")
            configured = switch.get("configured", False)
            role = switch.get("role", "Unknown")
            total_ports = switch.get("total_ports", 0)
            active_ports = switch.get("active_ports", 0)
            mtu = switch.get("mtu", "Unknown")
            port_speeds = switch.get("port_speeds", {})
            ports = switch.get("ports", [])

            # Capitalize switch type for display
            if switch_type.lower() == "cumulus":
                switch_type_display = "Cumulus Linux"
            else:
                switch_type_display = switch_type

            # Format configuration status
            config_status = "Configured" if configured else "Not Configured"

            # Switch header info
            switch_info_data = [
                ["Hostname", hostname],
                ["Model", model],
                ["Serial Number", serial],
                ["Firmware Version", firmware_version],
                ["Management IP", mgmt_ip],
                ["Switch Type", switch_type_display],
                ["State", state],
                ["Configuration Status", config_status],
                ["Role", role if role else "Not Assigned"],
                ["Total Ports", str(total_ports)],
                ["Active Ports", str(active_ports)],
                ["Port MTU", mtu],
            ]

            # Create heading and table elements
            switch_details_heading = self.brand_compliance.create_vast_section_heading(
                f"Switch {switch_num} Details: {switch_name}", level=2
            )
            switch_info_elements = self._create_cluster_info_table(switch_info_data, None)

            # Keep heading with first table (switch info)
            switch_section_elements = []
            switch_section_elements.extend(switch_details_heading)
            switch_section_elements.extend(switch_info_elements)
            content.append(KeepTogether(switch_section_elements))
            content.append(Spacer(1, 12))

            # Create port summary table with port numbers
            if ports:
                # Aggregate ports by speed, collecting port names
                port_summary: dict[str, Any] = {}

                for port in ports:
                    speed = port.get("speed")
                    if not speed or speed == "":
                        speed = "Unconfigured"

                    port_name = port.get("name", "Unknown")

                    if speed not in port_summary:
                        port_summary[speed] = []
                    port_summary[speed].append(port_name)

                # Create summary table with port count and port numbers
                summary_table_data = []
                headers = ["Port Count", "Speed", "Port Numbers"]

                # Sort by speed (200G, 100G, Unconfigured, then others)
                speed_order = {"200G": 0, "100G": 1, "Unconfigured": 2}
                sorted_summary = sorted(port_summary.items(), key=lambda x: speed_order.get(x[0], 3))

                for speed, port_list in sorted_summary:
                    # Sort port names naturally (swp1, swp2, ..., swp10, swp11, ...)
                    try:
                        sorted_ports = sorted(
                            port_list,
                            key=lambda x: (int("".join(filter(str.isdigit, x))) if any(c.isdigit() for c in x) else 0),
                        )
                    except Exception:
                        sorted_ports = sorted(port_list)

                    # Join ports with comma separation
                    ports_str = ", ".join(sorted_ports)

                    # Use Paragraph for port numbers to enable text wrapping
                    port_style = ParagraphStyle(
                        "PortNumbers",
                        parent=styles["Normal"],
                        fontSize=9,
                        alignment=0,  # Left alignment
                        wordWrap="CJK",  # Enable word wrapping
                    )
                    port_para = Paragraph(ports_str, port_style)

                    # Count ports for this speed
                    port_count = len(port_list)

                    summary_table_data.append([str(port_count), speed, port_para])

                # Create custom port summary table with specific column widths
                # Port Count: 15%, Speed: 15%, Port Numbers: 70%
                page_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
                col_widths = [
                    page_width * 0.15,  # Port Count (matches Speed column)
                    page_width * 0.15,  # Speed (reduced by 50%)
                    page_width * 0.70,  # Port Numbers (reduced slightly)
                ]

                # Prepare table data with headers
                full_table_data = [headers] + summary_table_data

                # Create table with custom column widths
                port_table = Table(full_table_data, colWidths=col_widths, repeatRows=1)

                # Apply VAST brand table styling
                table_style = TableStyle(
                    [
                        # Header row styling
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            self.brand_compliance.colors.BACKGROUND_DARK,
                        ),
                        (
                            "TEXTCOLOR",
                            (0, 0),
                            (-1, 0),
                            self.brand_compliance.colors.PURE_WHITE,
                        ),
                        (
                            "FONTNAME",
                            (0, 0),
                            (-1, 0),
                            self.brand_compliance.typography.PRIMARY_FONT,
                        ),
                        (
                            "FONTSIZE",
                            (0, 0),
                            (-1, 0),
                            self.brand_compliance.typography.BODY_SIZE,
                        ),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        # Data rows styling
                        (
                            "BACKGROUND",
                            (0, 1),
                            (-1, -1),
                            self.brand_compliance.colors.VAST_BLUE_LIGHTEST,
                        ),
                        (
                            "TEXTCOLOR",
                            (0, 1),
                            (-1, -1),
                            self.brand_compliance.colors.DARK_GRAY,
                        ),
                        (
                            "FONTNAME",
                            (0, 1),
                            (-1, -1),
                            self.brand_compliance.typography.BODY_FONT,
                        ),
                        (
                            "FONTSIZE",
                            (0, 1),
                            (-1, -1),
                            self.brand_compliance.typography.BODY_SIZE,
                        ),
                        # Borders and spacing
                        (
                            "GRID",
                            (0, 0),
                            (-1, -1),
                            1,
                            self.brand_compliance.colors.BACKGROUND_DARK,
                        ),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [
                                self.brand_compliance.colors.PURE_WHITE,
                                self.brand_compliance.colors.ALTERNATING_ROW,
                            ],
                        ),
                        ("PADDING", (0, 0), (-1, -1), 8),
                        ("LEFTPADDING", (0, 0), (-1, -1), 12),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                        ("WORDWRAP", (0, 0), (-1, -1), "CJK"),
                    ]
                )

                port_table.setStyle(table_style)

                # Create title and wrap with table in KeepTogether
                table_title = f"Port Summary: {switch_name}"
                title_para = Paragraph(
                    f"<b>{table_title}</b>",
                    self.brand_compliance.styles["vast_subheading"],
                )

                # Keep title and table together on page breaks
                port_summary_elements = [title_para, Spacer(1, 8), port_table]
                content.append(KeepTogether(port_summary_elements))
                content.append(Spacer(1, 12))

        return content

    def _create_logical_configuration(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """Create logical configuration section."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        content = []

        content.append(Paragraph("Logical Configuration", heading_style))
        content.append(Spacer(1, 12))

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Logical Configuration section documents the logical organization and data protection policies configured within the VAST Data cluster. This section provides visibility into tenant configurations, data views, access policies, VIP pools, and data protection settings including snapshot programs and protection policies. Understanding the logical configuration is crucial for data governance, access control validation, backup and recovery planning, and ensuring compliance with organizational data protection requirements. This information enables administrators to verify proper data isolation, validate backup schedules, and ensure that data protection policies align with business continuity objectives.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        sections = data.get("sections", {})
        logical_config = sections.get("logical_configuration", {}).get("data", {})

        # Prepare table data
        table_data = []

        # Tenants
        tenants = logical_config.get("tenants")
        if tenants:
            tenant_list = tenants.get("tenants", []) if isinstance(tenants, dict) else tenants
            tenant_count = len(tenant_list) if isinstance(tenant_list, list) else 0
            table_data.append(["Tenants", f"{tenant_count} tenants configured"])

        # Views
        views = logical_config.get("views")
        if views:
            view_list = views.get("views", []) if isinstance(views, dict) else views
            view_count = len(view_list) if isinstance(view_list, list) else 0
            table_data.append(["Views", f"{view_count} views configured"])

        # View Policies
        policies = logical_config.get("view_policies")
        if policies:
            policy_list = policies.get("policies", []) if isinstance(policies, dict) else policies
            policy_count = len(policy_list) if isinstance(policy_list, list) else 0
            table_data.append(["View Policies", f"{policy_count} policies configured"])

        # Protection Policies
        protection_policies = logical_config.get("protection_policies")
        if protection_policies:
            protection_policy_list = (
                protection_policies.get("policies", [])
                if isinstance(protection_policies, dict)
                else protection_policies
            )
            protection_policy_count = len(protection_policy_list) if isinstance(protection_policy_list, list) else 0
            table_data.append(
                [
                    "Protection Policies",
                    f"{protection_policy_count} policies configured",
                ]
            )

        # Network Services (VIP Pools, DNS and NTP from network configuration)
        network_config = sections.get("network_configuration", {}).get("data", {})
        if network_config:
            # VIP Pools
            vippools = network_config.get("vippools")
            if vippools:
                vippool_list = vippools.get("pools", []) if isinstance(vippools, dict) else vippools
                vippool_count = len(vippool_list) if isinstance(vippool_list, list) else 0
                table_data.append(["VIP Pools", f"{vippool_count} pools configured"])
            # DNS Configuration
            dns = network_config.get("dns")
            if dns:
                dns_list = dns.get("dns_servers", []) if isinstance(dns, dict) else dns
                if dns_list:
                    dns_servers = ", ".join(dns_list) if isinstance(dns_list, list) else str(dns_list)
                    table_data.append(["DNS Servers", dns_servers])

            # NTP Configuration
            ntp = network_config.get("ntp")
            if ntp:
                ntp_list = ntp.get("ntp_servers", []) if isinstance(ntp, dict) else ntp
                if ntp_list:
                    ntp_servers = ", ".join(ntp_list) if isinstance(ntp_list, list) else str(ntp_list)
                    table_data.append(["NTP Servers", ntp_servers])

        # Data Protection information
        protection_config = sections.get("data_protection_configuration", {}).get("data", {})

        # Snapshot Programs
        snapshots = protection_config.get("snapshot_programs")
        if snapshots:
            snapshot_list = snapshots.get("programs", []) if isinstance(snapshots, dict) else snapshots
            snapshot_count = len(snapshot_list) if isinstance(snapshot_list, list) else 0
            table_data.append(["Snapshot Programs", f"{snapshot_count} programs configured"])

        # Data Protection Protection Policies (from data protection configuration)
        data_protection_policies = protection_config.get("protection_policies")
        if data_protection_policies:
            data_protection_policy_list = (
                data_protection_policies.get("policies", [])
                if isinstance(data_protection_policies, dict)
                else data_protection_policies
            )
            data_protection_policy_count = (
                len(data_protection_policy_list) if isinstance(data_protection_policy_list, list) else 0
            )
            table_data.append(
                [
                    "Data Protection Policies",
                    f"{data_protection_policy_count} policies configured",
                ]
            )

        # Create table if we have data
        if table_data:
            table_elements = self.brand_compliance.create_vast_table(table_data, None, ["Resource", "Value"])
            content.extend(table_elements)
        else:
            # Fallback if no data
            content.append(
                Paragraph(
                    "No logical configuration data available.",
                    ParagraphStyle(
                        "Normal",
                        parent=styles["Normal"],
                        fontSize=self.config.font_size,
                    ),
                )
            )

        return content

    def _create_security_configuration(
        self,
        data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
    ) -> List[Any]:
        """Create security configuration section."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        content = []

        content.append(Paragraph("Security & Authentication", heading_style))
        content.append(Spacer(1, 12))

        # Place page marker immediately after heading to capture section start page
        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        # Section Overview
        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content.append(
            Paragraph(
                "The Security & Authentication section provides comprehensive documentation of all security-related configurations and authentication mechanisms implemented within the VAST Data cluster. This section covers authentication services including Active Directory, LDAP, and NIS integration, as well as security features such as data encryption settings, external key management (EKM) configuration, and security policy enforcement. Understanding the security configuration is essential for compliance auditing, security posture assessment, access control validation, and ensuring that the storage infrastructure meets organizational security requirements and industry best practices. This information supports security administrators in maintaining a robust security framework for the storage environment.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        sections = data.get("sections", {})
        security_config = sections.get("security_configuration", {}).get("data", {})

        # Prepare table data
        table_data = []

        # Active Directory
        ad_config = security_config.get("active_directory")
        if ad_config:
            table_data.append(
                [
                    "Authentication",
                    "Active Directory",
                    "Enabled",
                    str(ad_config.get("enabled", False)),
                ]
            )
            if ad_config.get("domain"):
                table_data.append(
                    [
                        "Authentication",
                        "Active Directory",
                        "Domain",
                        ad_config.get("domain"),
                    ]
                )
            servers = ad_config.get("servers", [])
            if servers:
                table_data.append(
                    [
                        "Authentication",
                        "Active Directory",
                        "Servers",
                        ", ".join(servers),
                    ]
                )

        # LDAP
        ldap_config = security_config.get("ldap")
        if ldap_config:
            table_data.append(
                [
                    "Authentication",
                    "LDAP",
                    "Enabled",
                    str(ldap_config.get("enabled", False)),
                ]
            )

        # NIS
        nis_config = security_config.get("nis")
        if nis_config:
            table_data.append(
                [
                    "Authentication",
                    "NIS",
                    "Enabled",
                    str(nis_config.get("enabled", False)),
                ]
            )

        # Encryption Configuration
        cluster_summary = data.get("cluster_summary", {})
        if cluster_summary:
            # Basic encryption settings
            enable_encryption = cluster_summary.get("enable_encryption")
            enable_encryption_display = enable_encryption if enable_encryption is not None else "Not Configured"
            table_data.append(["Security", "Encryption", "Enabled", str(enable_encryption_display)])

            encryption_type = cluster_summary.get("encryption_type")
            # Handle potential list values from API
            if isinstance(encryption_type, list):
                encryption_type_display = (
                    ", ".join(str(t) for t in encryption_type) if encryption_type else "Not Configured"
                )
            else:
                encryption_type_display = (
                    str(encryption_type) if encryption_type and encryption_type != "Unknown" else "Not Configured"
                )
            table_data.append(["Security", "Encryption", "Type", encryption_type_display])

            s3_aes_ciphers = cluster_summary.get("s3_enable_only_aes_ciphers")
            s3_aes_ciphers_display = s3_aes_ciphers if s3_aes_ciphers is not None else "Not Configured"
            table_data.append(
                [
                    "Security",
                    "Encryption",
                    "S3 AES Ciphers Only",
                    str(s3_aes_ciphers_display),
                ]
            )

            # External Key Management (EKM) settings
            ekm_servers = cluster_summary.get("ekm_servers")
            # Handle list values from API - convert to comma-separated string
            if isinstance(ekm_servers, list):
                ekm_servers_display = ", ".join(str(s) for s in ekm_servers) if ekm_servers else "Not Configured"
            else:
                ekm_servers_display = (
                    str(ekm_servers)
                    if ekm_servers and ekm_servers != "Unknown" and ekm_servers != ""
                    else "Not Configured"
                )
            table_data.append(["Security", "EKM", "Servers", ekm_servers_display])

            ekm_address = cluster_summary.get("ekm_address")
            # Handle list values from API - convert to comma-separated string
            if isinstance(ekm_address, list):
                ekm_address_display = ", ".join(str(a) for a in ekm_address) if ekm_address else "Not Configured"
            else:
                ekm_address_display = (
                    str(ekm_address)
                    if ekm_address and ekm_address != "Unknown" and ekm_address != ""
                    else "Not Configured"
                )
            table_data.append(["Security", "EKM", "Address", ekm_address_display])

            ekm_port = cluster_summary.get("ekm_port")
            ekm_port_display = ekm_port if ekm_port is not None else "Not Configured"
            table_data.append(["Security", "EKM", "Port", str(ekm_port_display)])

            ekm_auth_domain = cluster_summary.get("ekm_auth_domain")
            # Handle potential list values from API
            if isinstance(ekm_auth_domain, list):
                ekm_auth_domain_display = (
                    ", ".join(str(d) for d in ekm_auth_domain) if ekm_auth_domain else "Not Configured"
                )
            else:
                ekm_auth_domain_display = (
                    str(ekm_auth_domain)
                    if ekm_auth_domain and ekm_auth_domain != "Unknown" and ekm_auth_domain != ""
                    else "Not Configured"
                )
            table_data.append(["Security", "EKM", "Auth Domain", ekm_auth_domain_display])

            # Secondary EKM settings
            secondary_ekm_address = cluster_summary.get("secondary_ekm_address")
            # Handle potential list values from API
            if isinstance(secondary_ekm_address, list):
                secondary_ekm_address_display = (
                    ", ".join(str(a) for a in secondary_ekm_address) if secondary_ekm_address else "Not Configured"
                )
            else:
                secondary_ekm_address_display = (
                    str(secondary_ekm_address)
                    if secondary_ekm_address and secondary_ekm_address != "null"
                    else "Not Configured"
                )
            table_data.append(["Security", "Secondary EKM", "Address", secondary_ekm_address_display])

            secondary_ekm_port = cluster_summary.get("secondary_ekm_port")
            secondary_ekm_port_display = secondary_ekm_port if secondary_ekm_port is not None else "Not Configured"
            table_data.append(["Security", "Secondary EKM", "Port", str(secondary_ekm_port_display)])

        # Create table if we have data
        if table_data:
            table_elements = self.brand_compliance.create_vast_table(
                table_data,
                None,
                ["Type", "Description", "Function", "Value"],
            )
            content.extend(table_elements)
        else:
            # Fallback if no data
            content.append(
                Paragraph(
                    "No security configuration data available.",
                    ParagraphStyle(
                        "Normal",
                        parent=styles["Normal"],
                        fontSize=self.config.font_size,
                    ),
                )
            )

        return content

    # ------------------------------------------------------------------
    # Health Check & Post-Deployment Validation sections
    # ------------------------------------------------------------------

    @staticmethod
    def _fixup_health_results(
        health_data: Dict[str, Any],
        processed_data: Optional[Dict[str, Any]],
    ) -> None:
        """Correct stale health-check results from older JSON exports.

        Applies render-time patches so reports regenerated from pre-fix JSON
        files reflect current check logic without requiring a new live run.
        """
        results = health_data.get("results")
        if not results:
            return

        summary = health_data.get("summary")

        node_ssh = [r for r in results if r.get("category") == "node_ssh"]
        if node_ssh:
            kept = [r for r in results if r.get("category") != "node_ssh"]
            health_data["results"] = kept
            results = kept
            if summary:
                for r in node_ssh:
                    bucket = r.get("status", "error")
                    if bucket in summary:
                        summary[bucket] = max(0, summary[bucket] - 1)
                summary["total"] = sum(summary.get(k, 0) for k in ("pass", "fail", "warning", "skipped", "error"))

        cluster_summary = (processed_data or {}).get("cluster_summary", {})
        mgmt_cnode = cluster_summary.get("mgmt_cnode", "")

        hw = (processed_data or {}).get("hardware_inventory", {})
        cnodes_list = hw.get("cnodes") or []
        mgmt_names: set = set()
        if mgmt_cnode:
            mgmt_names.add(mgmt_cnode)
        for cn in cnodes_list:
            if cn.get("is_mgmt"):
                mgmt_names.add(cn.get("name", ""))

        for r in results:
            name = r.get("check_name", "")
            status = r.get("status", "")

            if name == "CNode Status" and status == "fail" and mgmt_names:
                msg = r.get("message", "")
                inactive = r.get("details", {}).get("inactive", [])
                only_mgmt = inactive and all(n in mgmt_names for n in inactive)
                disabled = r.get("details", {}).get("disabled", [])
                if only_mgmt and not disabled:
                    total = r.get("details", {}).get("total", len(inactive))
                    mgmt_label = inactive[0] if inactive else mgmt_cnode
                    r["status"] = "pass"
                    r["message"] = f"All {total} CNodes healthy (VMS on {mgmt_label})"
                    if summary:
                        summary["fail"] = max(0, summary.get("fail", 1) - 1)
                        summary["pass"] = summary.get("pass", 0) + 1

            elif name == "Active Alarms" and status == "fail":
                r["status"] = "warning"
                if summary:
                    summary["fail"] = max(0, summary.get("fail", 1) - 1)
                    summary["warning"] = summary.get("warning", 0) + 1

            elif name == "Switches in VMS" and status == "warning":
                msg = r.get("message", "")
                if "No switches registered" in msg or "no switches" in msg.lower():
                    r["status"] = "skipped"
                    r["message"] = "No switches registered"
                    if summary:
                        summary["warning"] = max(0, summary.get("warning", 1) - 1)
                        summary["skipped"] = summary.get("skipped", 0) + 1

            elif name == "Monitoring Config":
                r["status"] = "skipped"
                r["message"] = "Check removed — API endpoints unavailable"
                if summary and status != "skipped":
                    old_bucket = status if status in ("pass", "fail", "warning", "error") else "error"
                    summary[old_bucket] = max(0, summary.get(old_bucket, 1) - 1)
                    summary["skipped"] = summary.get("skipped", 0) + 1

            elif name == "Switch Config Backup":
                r["check_name"] = "Switch Config Readability"

            elif name == "VIP Pools" and status == "fail":
                msg = r.get("message", "")
                if "No VIP pools configured" in msg or "No enabled VIP pools" in msg:
                    r["status"] = "warning"
                    if summary:
                        summary["fail"] = max(0, summary.get("fail", 1) - 1)
                        summary["warning"] = summary.get("warning", 0) + 1

    def _create_health_check_section(
        self,
        health_data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
        processed_data: Optional[Dict[str, Any]] = None,
    ) -> List[Any]:
        """Create the Cluster Health Check Results section.

        Args:
            health_data: Health check results dict with 'summary' and 'results' keys.
            page_tracker: Optional page tracker for TOC page capture.
            section_key: Section key for PageMarker registration.
            processed_data: Full report data for render-time corrections on stale JSON.

        Returns:
            List of flowables for the section.
        """
        self._fixup_health_results(health_data, processed_data)
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content: List[Any] = []
        content.append(Paragraph("Cluster Health Check Results", heading_style))
        content.append(Spacer(1, 12))

        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        content.append(
            Paragraph(
                "This section presents the results of automated cluster health checks "
                "performed against the VAST Data cluster. Each check validates a specific "
                "aspect of cluster health including connectivity, service status, and "
                "configuration consistency. Results are summarised below with individual "
                "check details following the summary table.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        # Colour helpers for status cells
        status_colors = {
            "pass": self.brand_compliance.colors.SUCCESS_GREEN,
            "fail": self.brand_compliance.colors.ERROR_RED,
            "warning": self.brand_compliance.colors.WARNING_ORANGE,
            "skipped": self.brand_compliance.colors.MEDIUM_GRAY,
            "error": self.brand_compliance.colors.MEDIUM_GRAY,
        }

        # --- Summary table ---------------------------------------------------
        summary = health_data.get("summary", {})
        if summary:
            summary_headers = ["Pass", "Fail", "Warning", "Skipped", "Error", "Total"]
            summary_row = [
                self._safe_table_value(summary.get("pass", 0)),
                self._safe_table_value(summary.get("fail", 0)),
                self._safe_table_value(summary.get("warning", 0)),
                self._safe_table_value(summary.get("skipped", 0)),
                self._safe_table_value(summary.get("error", 0)),
                self._safe_table_value(summary.get("total", 0)),
            ]

            page_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
            col_width = page_width / len(summary_headers)
            summary_table = Table(
                [summary_headers, summary_row],
                colWidths=[col_width] * len(summary_headers),
                repeatRows=1,
            )

            cell_styles = [
                ("BACKGROUND", (0, 0), (-1, 0), self.brand_compliance.colors.BACKGROUND_DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), self.brand_compliance.colors.PURE_WHITE),
                ("FONTNAME", (0, 0), (-1, 0), self._font("bold")),
                ("FONTSIZE", (0, 0), (-1, -1), self.config.font_size),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 1, self.brand_compliance.colors.BACKGROUND_DARK),
                ("PADDING", (0, 0), (-1, -1), 8),
                # Per-column colour coding for data row
                ("BACKGROUND", (0, 1), (0, 1), status_colors["pass"]),
                ("TEXTCOLOR", (0, 1), (0, 1), self.brand_compliance.colors.PURE_WHITE),
                ("BACKGROUND", (1, 1), (1, 1), status_colors["fail"]),
                ("TEXTCOLOR", (1, 1), (1, 1), self.brand_compliance.colors.PURE_WHITE),
                ("BACKGROUND", (2, 1), (2, 1), status_colors["warning"]),
                ("TEXTCOLOR", (2, 1), (2, 1), self.brand_compliance.colors.DARK_GRAY),
                ("BACKGROUND", (3, 1), (3, 1), status_colors["skipped"]),
                ("TEXTCOLOR", (3, 1), (3, 1), self.brand_compliance.colors.PURE_WHITE),
                ("BACKGROUND", (4, 1), (4, 1), status_colors["error"]),
                ("TEXTCOLOR", (4, 1), (4, 1), self.brand_compliance.colors.PURE_WHITE),
                ("BACKGROUND", (5, 1), (5, 1), self.brand_compliance.colors.BACKGROUND_DARK),
                ("TEXTCOLOR", (5, 1), (5, 1), self.brand_compliance.colors.PURE_WHITE),
            ]

            summary_table.setStyle(TableStyle(cell_styles))
            content.append(Paragraph("<b>Health Check Summary</b>", styles["Normal"]))
            content.append(Spacer(1, 4))
            content.append(summary_table)
            content.append(Spacer(1, 16))

        # --- Detailed results table -------------------------------------------
        results = health_data.get("results", [])
        if results:
            detail_headers = ["Check Name", "Category", "Status", "Message"]
            detail_data = [detail_headers]

            cell_font = self._font("regular")
            cell_size = self.config.font_size - 1
            cell_style = ParagraphStyle(
                "HealthCell",
                parent=styles["Normal"],
                fontName=cell_font,
                fontSize=cell_size,
                leading=cell_size + 2,
                alignment=1,
                wordWrap="CJK",
            )
            msg_style = ParagraphStyle(
                "HealthMsg",
                parent=cell_style,
                alignment=0,
            )

            for r in results:
                detail_data.append(
                    [
                        Paragraph(self._safe_table_value(r.get("check_name")), cell_style),
                        Paragraph(self._safe_table_value(r.get("category")), cell_style),
                        Paragraph(self._safe_table_value(r.get("status")), cell_style),
                        Paragraph(self._safe_table_value(r.get("message")), msg_style),
                    ]
                )

            page_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
            col_widths = [
                page_width * 0.22,  # Check Name
                page_width * 0.12,  # Category
                page_width * 0.10,  # Status
                page_width * 0.56,  # Message
            ]

            detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)

            detail_styles = [
                ("BACKGROUND", (0, 0), (-1, 0), self.brand_compliance.colors.BACKGROUND_DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), self.brand_compliance.colors.PURE_WHITE),
                ("FONTNAME", (0, 0), (-1, 0), self._font("bold")),
                ("FONTSIZE", (0, 0), (-1, 0), self.config.font_size - 1),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 1, self.brand_compliance.colors.BACKGROUND_DARK),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [self.brand_compliance.colors.PURE_WHITE, self.brand_compliance.colors.ALTERNATING_ROW],
                ),
            ]

            # Per-row status colour in the Status column (index 2)
            for row_idx, r in enumerate(results, start=1):
                status_key = str(r.get("status", "")).lower()
                bg = status_colors.get(status_key, self.brand_compliance.colors.ALTERNATING_ROW)
                fg = (
                    self.brand_compliance.colors.DARK_GRAY
                    if status_key == "warning"
                    else self.brand_compliance.colors.PURE_WHITE
                )
                detail_styles.append(("BACKGROUND", (2, row_idx), (2, row_idx), bg))
                detail_styles.append(("TEXTCOLOR", (2, row_idx), (2, row_idx), fg))

            detail_table.setStyle(TableStyle(detail_styles))
            content.append(Paragraph("<b>Detailed Check Results</b>", styles["Normal"]))
            content.append(Spacer(1, 4))
            content.append(detail_table)
            content.append(Spacer(1, 12))
        elif not summary:
            content.append(
                Paragraph(
                    "No health check data available.",
                    ParagraphStyle("Normal", parent=styles["Normal"], fontSize=self.config.font_size),
                )
            )

        return content

    def _create_post_deployment_activities_section(
        self,
        activities_data: Dict[str, Any],
        page_tracker: Optional[Dict[str, int]] = None,
        section_key: Optional[str] = None,
        processed_data: Optional[Dict[str, Any]] = None,
    ) -> List[Any]:
        """Create the Post Deployment Activities section with next-steps checklist.

        Args:
            activities_data: Dict with 'next_steps' list of {item, description, status} dicts.
            page_tracker: Optional page tracker for TOC page capture.
            section_key: Section key for PageMarker registration.
            processed_data: Full report data for render-time status resolution fallback.

        Returns:
            List of flowables for the section.
        """
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        overview_style = ParagraphStyle(
            "Section_Overview",
            parent=styles["Normal"],
            fontSize=self.config.font_size - 1,
            textColor=self.brand_compliance.colors.BACKGROUND_DARK,
            spaceAfter=12,
            spaceBefore=8,
            leftIndent=12,
            rightIndent=12,
        )

        content: List[Any] = []
        content.append(Paragraph("Post Deployment Activities", heading_style))
        content.append(Spacer(1, 12))

        if page_tracker is not None and section_key:
            content.append(PageMarker(section_key, page_tracker))

        content.append(
            Paragraph(
                "This section outlines the recommended next steps to complete after "
                "cluster installation and validation. Each item should be addressed "
                "before handing the cluster over to the customer for production use. "
                "Refer to the VAST Installation Template for detailed procedures.",
                overview_style,
            )
        )
        content.append(Spacer(1, 8))

        next_steps = activities_data.get("next_steps", [])

        if next_steps and not next_steps[0].get("status"):
            self._resolve_post_deploy_status_at_render(next_steps, processed_data or {})

        if next_steps:
            cell_font_size = self.config.font_size - 1

            cell_style = ParagraphStyle(
                "ChecklistCell",
                parent=styles["Normal"],
                fontSize=cell_font_size,
                leading=cell_font_size + 2,
                wordWrap="CJK",
            )
            cell_style_bold = ParagraphStyle(
                "ChecklistCellBold",
                parent=cell_style,
                fontName=self._font("bold"),
            )
            cell_style_center = ParagraphStyle(
                "ChecklistCellCenter",
                parent=cell_style,
                alignment=1,
            )
            header_style_cell = ParagraphStyle(
                "ChecklistHeader",
                parent=cell_style,
                fontName=self._font("bold"),
                textColor=self.brand_compliance.colors.PURE_WHITE,
            )

            checklist_data = [
                [
                    Paragraph("Item", header_style_cell),
                    Paragraph("Description", header_style_cell),
                    Paragraph("Status", header_style_cell),
                ]
            ]

            status_colors = {
                "Completed": self.brand_compliance.colors.SUCCESS_GREEN,
                "Optional": self.brand_compliance.colors.ACCENT_BLUE,
                "Pending": self.brand_compliance.colors.WARNING_ORANGE,
            }

            for step in next_steps:
                step_status = step.get("status", "Pending")
                checklist_data.append(
                    [
                        Paragraph(self._safe_table_value(step.get("item")), cell_style_bold),
                        Paragraph(self._safe_table_value(step.get("description")), cell_style),
                        Paragraph(step_status, cell_style_center),
                    ]
                )

            page_width = getattr(self, "_frame_width", A4[0] - 1.0 * inch)
            col_widths = [page_width * 0.25, page_width * 0.55, page_width * 0.20]

            checklist_table = Table(checklist_data, colWidths=col_widths, repeatRows=1)

            checklist_styles = [
                ("BACKGROUND", (0, 0), (-1, 0), self.brand_compliance.colors.BACKGROUND_DARK),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 1, self.brand_compliance.colors.BACKGROUND_DARK),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [self.brand_compliance.colors.PURE_WHITE, self.brand_compliance.colors.ALTERNATING_ROW],
                ),
            ]

            for row_idx, step in enumerate(next_steps, start=1):
                step_status = step.get("status", "Pending")
                bg = status_colors.get(step_status, self.brand_compliance.colors.WARNING_ORANGE)
                text_color = (
                    self.brand_compliance.colors.PURE_WHITE
                    if step_status in ("Completed", "Optional")
                    else self.brand_compliance.colors.DARK_GRAY
                )
                checklist_styles.append(("BACKGROUND", (2, row_idx), (2, row_idx), bg))
                checklist_styles.append(("TEXTCOLOR", (2, row_idx), (2, row_idx), text_color))

            checklist_table.setStyle(TableStyle(checklist_styles))
            content.append(Paragraph("<b>Next Steps — Get Started Using VAST Data</b>", styles["Normal"]))
            content.append(Spacer(1, 4))
            content.append(checklist_table)
            content.append(Spacer(1, 12))
        else:
            content.append(
                Paragraph(
                    "No post-deployment activity items available.",
                    ParagraphStyle("Normal", parent=styles["Normal"], fontSize=self.config.font_size),
                )
            )

        return content

    @staticmethod
    def _resolve_post_deploy_status_at_render(next_steps: List[Dict[str, Any]], processed_data: Dict[str, Any]) -> None:
        """Render-time fallback: resolve post-deploy status from health check data in the report JSON."""
        hc_results: Dict[str, str] = {}
        hc_section = processed_data.get("sections", {}).get("health_check", {})
        for r in hc_section.get("data", {}).get("results", []):
            hc_results[r.get("check_name", "")] = r.get("status", "")

        cluster = processed_data.get("cluster_summary", {})
        license_val = str(cluster.get("license", "") or "").strip().lower()

        auto_checks: Dict[str, bool] = {
            "Configure Call Home w/ Cloud Integration": hc_results.get("Call Home Status") == "pass",
            "Create VIP": hc_results.get("VIP Pools") == "pass",
            "Activate License": (
                hc_results.get("License") == "pass"
                or (bool(license_val) and license_val not in ("", "none", "unknown"))
            ),
        }
        manual_items = {
            "Test Fail-over Behavior",
            "Confirm VIP Movement and ARP Updates",
            "Change Default Passwords",
        }

        for step in next_steps:
            item = step.get("item", "")
            if item in manual_items:
                step["status"] = "Optional"
            elif item in auto_checks:
                step["status"] = "Completed" if auto_checks[item] else "Pending"
            else:
                step["status"] = "Pending"

    def _create_data_protection_configuration(self, data: Dict[str, Any]) -> List[Any]:
        """Create data protection configuration section."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        content.append(Paragraph("Data Protection", heading_style))
        content.append(Spacer(1, 12))

        sections = data.get("sections", {})
        protection_config = sections.get("data_protection_configuration", {}).get("data", {})

        # Snapshot Programs
        snapshots = protection_config.get("snapshot_programs")
        if snapshots:
            content.append(Paragraph("<b>Snapshot Programs:</b>", normal_style))
            for snapshot in snapshots.get("programs", []):
                content.append(
                    Paragraph(
                        f"• {snapshot.get('name', 'Unknown')} - {snapshot.get('schedule', 'Unknown')} ({'Enabled' if snapshot.get('enabled') else 'Disabled'})",
                        normal_style,
                    )
                )
            content.append(Spacer(1, 8))

        # Protection Policies
        policies = protection_config.get("protection_policies")
        if policies:
            policy_list = policies.get("policies", []) if isinstance(policies, dict) else policies
            policy_count = len(policy_list) if isinstance(policy_list, list) else 0
            content.append(
                Paragraph(
                    f"<b>Protection Policies:</b> {policy_count} policies configured",
                    normal_style,
                )
            )

        return content

    def _create_enhanced_features_section(self, data: Dict[str, Any]) -> List[Any]:
        """Create enhanced features section."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        content.append(Paragraph("Enhanced Features", heading_style))
        content.append(Spacer(1, 12))

        enhanced_features = data.get("metadata", {}).get("enhanced_features", {})

        # Rack Height Support
        rack_support = enhanced_features.get("rack_height_supported", False)
        content.append(Paragraph("<b>Rack Positioning:</b>", normal_style))
        content.append(Paragraph(f"• Supported: {'Yes' if rack_support else 'No'}", normal_style))
        if rack_support:
            content.append(
                Paragraph(
                    "• Automated U-number generation for hardware positioning",
                    normal_style,
                )
            )
            content.append(Paragraph("• Physical rack layout visualization available", normal_style))
        else:
            content.append(Paragraph("• Manual entry required for rack positions", normal_style))
        content.append(Spacer(1, 8))

        # PSNT Support
        psnt_support = enhanced_features.get("psnt_supported", False)
        content.append(Paragraph("<b>PSNT Tracking:</b>", normal_style))
        content.append(Paragraph(f"• Supported: {'Yes' if psnt_support else 'No'}", normal_style))
        if psnt_support:
            content.append(
                Paragraph(
                    "• Cluster Product Serial Number available for support tracking",
                    normal_style,
                )
            )
            cluster_info = data.get("cluster_summary", {})
            psnt = cluster_info.get("psnt")
            if psnt:
                content.append(Paragraph(f"• PSNT: {psnt}", normal_style))
        else:
            content.append(Paragraph("• PSNT not available for this cluster version", normal_style))

        return content

    def _create_appendix(self, data: Dict[str, Any]) -> List[Any]:
        """Create appendix section."""
        styles = getSampleStyleSheet()

        heading_style = ParagraphStyle(
            "Section_Heading",
            parent=styles["Heading1"],
            fontSize=self.config.heading_font_size,
            spaceAfter=12,
        )

        normal_style = ParagraphStyle(
            "Section_Normal",
            parent=styles["Normal"],
            fontSize=self.config.font_size,
            spaceAfter=8,
        )

        content = []

        content.append(Paragraph("Appendix", heading_style))
        content.append(Spacer(1, 12))

        # Generation metadata
        metadata = data.get("metadata", {})
        content.append(Paragraph("<b>Report Generation Information:</b>", normal_style))
        content.append(
            Paragraph(
                f"• Generated on: {metadata.get('extraction_timestamp', 'Unknown')}",
                normal_style,
            )
        )
        content.append(
            Paragraph(
                f"• Data completeness: {metadata.get('overall_completeness', 0.0):.1%}",
                normal_style,
            )
        )
        content.append(Paragraph(f"• API version: {metadata.get('api_version', 'Unknown')}", normal_style))
        content.append(
            Paragraph(
                f"• Cluster version: {metadata.get('cluster_version', 'Unknown')}",
                normal_style,
            )
        )
        content.append(Spacer(1, 12))

        # Physical layout information
        hardware = data.get("hardware_inventory", {})
        physical_layout = hardware.get("physical_layout")
        if physical_layout:
            content.append(Paragraph("<b>Physical Rack Layout:</b>", normal_style))
            stats = physical_layout.get("statistics", {})
            content.append(
                Paragraph(
                    f"• Occupied positions: {stats.get('occupied_positions', 0)}",
                    normal_style,
                )
            )
            content.append(
                Paragraph(
                    f"• Position range: U{stats.get('min_position', 0)} - U{stats.get('max_position', 0)}",
                    normal_style,
                )
            )
            content.append(Paragraph(f"• Total CNodes: {stats.get('total_cnodes', 0)}", normal_style))
            content.append(Paragraph(f"• Total DNodes: {stats.get('total_dnodes', 0)}", normal_style))

        return content


# Convenience function for easy usage
def create_report_builder(
    config: Optional[ReportConfig] = None,
    library_path: Optional[str] = None,
    user_images_dir: Optional[str] = None,
) -> VastReportBuilder:
    """
    Create and return a configured VastReportBuilder instance.

    Args:
        config (ReportConfig, optional): Report configuration
        library_path: Path to user device_library.json
        user_images_dir: Path to user-uploaded hardware images directory

    Returns:
        VastReportBuilder: Configured report builder instance
    """
    return VastReportBuilder(config, library_path, user_images_dir)


if __name__ == "__main__":
    """
    Test the report builder when run as a standalone module.
    """
    from utils.logger import setup_logging

    # Set up logging
    setup_logging()
    logger = get_logger(__name__)

    logger.info("VAST Report Builder Module Test")
    logger.info("This module generates professional PDF reports from processed data")
    logger.info("Enhanced features: rack positioning and PSNT integration")
    logger.info("Ready for integration with main CLI application")
